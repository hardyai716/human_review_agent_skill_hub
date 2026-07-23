#!/usr/bin/env python3
"""Build reusable label-rate notification artifacts and optional Card pushes."""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from card_hash import strip_internal_keys, verify_card_hash
from render_label_rate_grading_card import (
    LEVEL_COLORS,
    card_design_check,
    render_grading_card,
)
from resolve_label_rate_poc_routing import (
    build_poc_routing_plan,
    load_poc_mapping,
    load_stage_1_sample,
    poc_mapping_index,
    resolve_row_poc,
)
from sheet_importer import import_xlsx_as_feishu_sheet, run_lark_cli


ROOT = Path(__file__).resolve().parents[3]
SCENARIO_KEY = "efficiency-label-rate"
SCENARIO_REFERENCE = "references/scenarios/efficiency-label-rate.md"
CARD_TEMPLATE_ASSET = (
    "assets/efficiency-label-rate/low_efficiency_grading_card_template.json"
)
CARD_SCHEMA_NOTES_ASSET = "assets/efficiency-label-rate/card_schema_notes.md"
POC_MAPPING_ASSET = "assets/efficiency-label-rate/mach_root_label_poc_mapping.json"
REPORT_TYPE = "low_efficiency_grading"
CSV_LEVELS = ["notice", "P2", "P1", "P0"]
CARD_LEVELS = ["P0", "P1", "P2", "notice"]
FILTERED_COMPREHENSIVE_CSV = "综合_剔除+1同意.csv"
FILTERED_SUMMARY_CSV = "汇总统计_剔除+1同意.csv"
LEVEL_COLUMN_SPECS = [
    ("data_source", "数据来源"),
    ("warning_dimension", "预警维度"),
    ("severity_level", "预警等级"),
    ("mach_root_label_name", "机审一级标签"),
    ("strategy_id", "策略ID"),
    ("strategy_name", "策略名称"),
    ("max_data_date", "最大有数日期"),
    ("POC", "POC"),
    ("avg_review_in_cnt", "日均进审量"),
    ("avg_review_done_cnt", "日均完审量"),
    ("avg_label_cnt", "日均打标量"),
    ("label_rate", "打标率"),
    ("hit_conditions", "命中原因"),
    ("is_plus1_agreed", "是否+1同意"),
    ("plus1_update_date", "更新日期"),
    (
        "plus1_agreed_before_current_period",
        "+1同意日期是否在本次统计周期前",
    ),
]
SUMMARY_COLUMN_SPECS = [
    ("data_source", "数据来源"),
    ("mach_root_label_name", "机审一级标签"),
    ("POC", "POC"),
    ("low_efficiency_strategy_count", "低效策略数"),
    ("avg_review_in_cnt", "低效策略日均进审量"),
    ("avg_review_done_cnt", "低效策略日均完审量"),
    ("avg_label_cnt", "低效策略日均打标量"),
    ("label_rate", "低效策略打标率"),
]


@dataclass(frozen=True)
class ReportArtifacts:
    """CSV/XLSX report files and the rows used by card summary tables."""

    summary_rows: list[dict[str, Any]]
    filtered_summary_rows: list[dict[str, Any]]
    workbook_path: Path
    summary_csv_path: Path
    filtered_summary_csv_path: Path
    filtered_comprehensive_csv_path: Path
    filtered_comprehensive_row_count: int


@dataclass(frozen=True)
class CardArtifacts:
    """Card payload paths and hash-check metadata."""

    card_path: Path
    card_with_meta_path: Path
    hash_check_path: Path
    card_json: dict[str, Any]
    card_with_meta: dict[str, Any]
    hash_check: dict[str, Any]


@dataclass(frozen=True)
class NotificationArtifacts:
    """Top-level notification artifact paths."""

    output_dir: Path
    summary_path: Path
    notification_draft_path: Path
    send_plan_path: Path
    poc_routing_path: Path
    publish_summary_path: Path
    report: ReportArtifacts
    card: CardArtifacts
    summary: dict[str, Any]
    notification_draft: dict[str, Any]
    send_plan: dict[str, Any]
    publish_summary: dict[str, Any]


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Build notification_draft, send_plan, CSV/XLSX reports, POC routing, "
            "and Card JSON for a label-rate grading analysis_result JSONL. "
            "When --send-user-id is provided, the generated Feishu interactive "
            "Card JSON is sent to that user."
        )
    )
    parser.add_argument("--source", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--top-n", type=int, default=10)
    parser.add_argument("--sheet-url")
    parser.add_argument(
        "--import-sheet",
        action="store_true",
        help=(
            "Opt in to importing the XLSX report as a Feishu online sheet "
            "(a real online write). Off by default; debug_only stays local."
        ),
    )
    parser.add_argument("--identity", choices=["bot", "user"], default="bot")
    parser.add_argument("--title", default="近7天低效打标策略全等级结果")
    parser.add_argument(
        "--send-user-id",
        help=(
            "Feishu open_id for a direct-message push. Sends the generated "
            "interactive Card JSON; does not downgrade to text/markdown."
        ),
    )
    parser.add_argument("--idempotency-key")
    args = parser.parse_args()

    if args.top_n <= 0:
        raise SystemExit("--top-n must be a positive integer.")

    artifacts = build_label_rate_notification_artifacts(
        source_path=Path(args.source),
        output_dir=Path(args.output_dir),
        top_n=args.top_n,
        sheet_url=args.sheet_url,
        identity=args.identity,
        title=args.title,
        self_send_requested=bool(args.send_user_id),
        sent_payload=None,
        target_user_id=args.send_user_id,
        target_chat_id=None,
        auto_import_sheet=args.import_sheet,
    )

    sent_payload = send_card_if_requested(args, artifacts.card.card_path)
    if sent_payload is not None:
        write_json(Path(args.output_dir) / "lark_send_result.json", sent_payload)
        sheet_url = artifacts.summary.get("sheet_url") or args.sheet_url
        artifacts = build_label_rate_notification_artifacts(
            source_path=Path(args.source),
            output_dir=Path(args.output_dir),
            top_n=args.top_n,
            sheet_url=sheet_url,
            identity=args.identity,
            title=args.title,
            self_send_requested=True,
            sent_payload=sent_payload,
            target_user_id=args.send_user_id,
            target_chat_id=None,
            auto_import_sheet=args.import_sheet,
        )
    print(
        "Label-rate notification artifacts wrote "
        f"{relative_to_root(artifacts.output_dir)}; sent="
        f"{artifacts.publish_summary['sent']}; message_id="
        f"{artifacts.publish_summary['message_id']}"
    )


def send_card_if_requested(
    args: argparse.Namespace,
    card_path: Path,
) -> dict[str, Any] | None:
    if not args.send_user_id:
        return None
    idempotency_key = args.idempotency_key or safe_idempotency_key(
        f"{REPORT_TYPE}-{Path(args.output_dir).name}-{args.send_user_id}"
    )
    return send_interactive_card(
        card_path=card_path,
        user_id=args.send_user_id,
        identity=args.identity,
        idempotency_key=idempotency_key,
    )


def send_interactive_card(
    *,
    card_path: Path,
    user_id: str,
    identity: str,
    idempotency_key: str,
) -> dict[str, Any]:
    payload = run_lark_cli(
        [
            "lark-cli",
            "im",
            "+messages-send",
            "--json",
            "--as",
            identity,
            "--user-id",
            user_id,
            "--msg-type",
            "interactive",
            "--content",
            card_path.read_text(encoding="utf-8"),
            "--idempotency-key",
            idempotency_key,
        ]
    )
    if payload.get("ok") is not True:
        raise RuntimeError(f"Feishu Card send failed: {payload}")
    return payload


def build_label_rate_notification_artifacts(
    *,
    source_path: Path,
    output_dir: Path,
    top_n: int,
    sheet_url: str | None,
    identity: str,
    title: str,
    self_send_requested: bool,
    sent_payload: dict[str, Any] | None,
    target_user_id: str | None,
    target_chat_id: str | None,
    auto_import_sheet: bool = False,
) -> NotificationArtifacts:
    """Build all safe notification artifacts; caller owns any real send action."""

    if top_n <= 0:
        raise ValueError("top_n must be a positive integer.")
    output_dir.mkdir(parents=True, exist_ok=True)
    publish_dir = output_dir / "publish"
    publish_dir.mkdir(parents=True, exist_ok=True)

    sample = adapt_analysis_sample(load_stage_1_sample(source_path))
    execution = sample["readonly_execution"]
    period = derive_period(sample)
    report = write_report_artifacts(output_dir, execution, period)
    online_write_attempted = False
    online_write_executed = False
    if auto_import_sheet and not sheet_url:
        online_write_attempted = True
        sheet_url = import_xlsx_as_feishu_sheet(
            workbook_path=report.workbook_path,
            output_dir=output_dir,
            sheet_name=(
                f"低效打标全等级结果-{period['current_start']}-"
                f"{period['current_end']}"
            ),
        )
        online_write_executed = bool(sheet_url)
    summary = build_summary(
        source_path=source_path,
        output_dir=output_dir,
        workbook_path=report.workbook_path,
        sample=sample,
        period=period,
        sheet_url=sheet_url,
        top_n=top_n,
        self_send_requested=self_send_requested,
        label_poc_summary_count=len(report.summary_rows),
        filtered_label_poc_summary_count=len(report.filtered_summary_rows),
        filtered_comprehensive_row_count=report.filtered_comprehensive_row_count,
        online_write_attempted=online_write_attempted,
        online_write_executed=online_write_executed,
    )
    poc_routing_path = write_poc_routing_artifact(
        output_dir=output_dir,
        sample=sample,
        source_path=source_path,
        online_write_executed=online_write_executed,
    )
    card = write_card_artifacts(
        publish_dir=publish_dir,
        summary=summary,
        summary_rows=report.summary_rows,
        level_results=execution["level_results"],
        top_n=top_n,
        sheet_url=sheet_url,
        title=title,
        full_evidence_rows=execution["comprehensive_results"],
    )

    summary_path = output_dir / "summary.json"
    notification_draft_path = output_dir / "notification_draft.json"
    send_plan_path = output_dir / "send_plan.json"
    publish_summary_path = publish_dir / f"{REPORT_TYPE}.publish_summary.json"
    publish_summary = build_publish_summary(
        source_path=source_path,
        output_dir=output_dir,
        summary_path=summary_path,
        workbook_path=report.workbook_path,
        summary_csv_path=report.summary_csv_path,
        filtered_summary_csv_path=report.filtered_summary_csv_path,
        sheet_url=sheet_url,
        card=card,
        notification_draft_path=notification_draft_path,
        send_plan_path=send_plan_path,
        identity=identity,
        sent_payload=sent_payload,
        target_user_id=target_user_id,
        target_chat_id=target_chat_id,
        online_write_attempted=online_write_attempted,
        online_write_executed=online_write_executed,
    )
    summary["publish"] = publish_summary
    notification_draft = build_notification_draft(
        summary=summary,
        poc_routing_path=poc_routing_path,
        card_path=card.card_path,
        card_with_meta_path=card.card_with_meta_path,
        hash_check_path=card.hash_check_path,
        publish_summary=publish_summary,
    )
    send_plan = build_send_plan(
        identity=identity,
        publish_summary=publish_summary,
        poc_routing_path=poc_routing_path,
        card_path=card.card_path,
        notification_draft_path=notification_draft_path,
    )

    write_json(notification_draft_path, notification_draft)
    write_json(send_plan_path, send_plan)
    write_json(summary_path, summary)
    write_json(publish_summary_path, publish_summary)
    return NotificationArtifacts(
        output_dir=output_dir,
        summary_path=summary_path,
        notification_draft_path=notification_draft_path,
        send_plan_path=send_plan_path,
        poc_routing_path=poc_routing_path,
        publish_summary_path=publish_summary_path,
        report=report,
        card=card,
        summary=summary,
        notification_draft=notification_draft,
        send_plan=send_plan,
        publish_summary=publish_summary,
    )


def adapt_analysis_sample(sample: dict[str, Any]) -> dict[str, Any]:
    """Normalize source profiles to the notification evidence contract."""

    if sample.get("analysis_mode") != "report_flow_low_label_rate":
        return sample

    adapted = dict(sample)
    execution = dict(sample["readonly_execution"])
    source_rows = [
        dict(zip(execution.get("columns", []), row))
        if isinstance(row, list)
        else row
        for row in execution.get("rows", [])
    ]
    routing_index = poc_mapping_index(load_poc_mapping())
    notice_rows = [
        normalize_report_flow_row(row, index, routing_index)
        for index, row in enumerate(source_rows)
    ]
    level_results = {
        "notice": {
            "severity_level": "notice",
            "severity_priority": 3,
            "row_count": len(notice_rows),
            "source_row_count": len(notice_rows),
            "truncated": execution.get("truncated"),
            "columns": list(notice_rows[0]) if notice_rows else [],
            "rows": notice_rows,
        },
        "P2": empty_level_result("P2", 2),
        "P1": empty_level_result("P1", 1),
        "P0": empty_level_result("P0", 0),
    }
    execution.update(
        {
            "level_counts": {
                "notice": len(notice_rows),
                "P2": 0,
                "P1": 0,
                "P0": 0,
            },
            "level_results": level_results,
            "comprehensive_results": notice_rows,
            "row_count": len(notice_rows),
            "evidence_fields": [
                "data_direction",
                "enpool_reason",
                "avg_report_review_done_cnt",
                "avg_report_label_cnt",
                "report_label_rate",
            ],
        }
    )
    adapted["readonly_execution"] = execution
    adapted["data_direction"] = "report_flow"
    adapted["source_profile"] = "report_flow_review"
    return adapted


def empty_level_result(level: str, priority: int) -> dict[str, Any]:
    return {
        "severity_level": level,
        "severity_priority": priority,
        "row_count": 0,
        "source_row_count": 0,
        "truncated": False,
        "columns": [],
        "rows": [],
    }


def normalize_report_flow_row(
    row: dict[str, Any],
    index: int,
    routing_index: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    avg_done = float(row.get("avg_report_review_done_cnt", 0) or 0)
    avg_label = float(row.get("avg_report_label_cnt", 0) or 0)
    rate = float(row.get("report_label_rate", 0) or 0)
    reason = str(row.get("enpool_reason") or "（空/enpool_reason）")
    routing = resolve_row_poc(
        {
            **row,
            "data_direction": "report_flow",
            "enpool_reason": reason,
        },
        routing_index,
    )
    poc_name = routing.get("poc_name") or "未映射"
    return {
        **row,
        "data_direction": "report_flow",
        "data_source": "举报流转",
        "warning_dimension": "举报原因维度",
        "severity_level": "notice",
        "severity_priority": 3,
        "mach_root_label_name": "举报",
        "strategy_id": reason,
        "strategy_name": reason,
        "enpool_reason": reason,
        "POC": poc_name,
        "poc_name": poc_name,
        "poc_open_id": routing.get("poc_open_id"),
        "poc_mapping_status": routing.get("mapping_status"),
        "max_data_date": row.get("max_data_date", ""),
        "avg_review_in_cnt": avg_done,
        "avg_review_done_cnt": avg_done,
        "avg_label_cnt": avg_label,
        "label_rate": rate,
        "is_plus1_agreed": "否",
        "plus1_update_date": "",
        "plus1_agreed_before_current_period": "否",
        "hit_rule_id": f"report_flow_low_label_rate_{index}",
        "hit_rule_ids": [f"report_flow_low_label_rate_{index}"],
        "hit_condition": "举报打标率<10%且人审完结量>0",
        "hit_conditions": ["举报打标率<10%且人审完结量>0"],
    }


def derive_period(sample: dict[str, Any]) -> dict[str, str]:
    time_range = sample.get("QueryPlan", {}).get("time_range", {})
    if (
        isinstance(time_range, dict)
        and time_range.get("current_start")
        and time_range.get("current_end")
    ):
        current_start = str(time_range["current_start"])
        current_end = str(time_range["current_end"])
        history_start = str(time_range.get("history_start") or current_start)
        checked_at = str(time_range.get("checked_at") or "")
        if not checked_at:
            checked_at = datetime.combine(
                datetime.fromisoformat(current_end).date() + timedelta(days=1),
                datetime.min.time(),
            ).isoformat()
        return {
            "current_start": current_start,
            "current_end": current_end,
            "history_start": history_start,
            "checked_at": checked_at,
        }

    freshness = sample.get("source_footer", {}).get("data_freshness", "")
    match = re.search(r"checked_at=([^;\\s]+)", freshness)
    if match:
        checked_at = datetime.fromisoformat(match.group(1))
    else:
        checked_at = datetime.now()
    current_end = checked_at.date() - timedelta(days=1)
    current_start = current_end - timedelta(days=6)
    history_start = current_end - timedelta(days=27)
    return {
        "current_start": current_start.isoformat(),
        "current_end": current_end.isoformat(),
        "history_start": history_start.isoformat(),
        "checked_at": checked_at.isoformat(),
    }


def write_report_artifacts(
    output_dir: Path,
    execution: dict[str, Any],
    period: dict[str, str],
) -> ReportArtifacts:
    current_start = period["current_start"]
    level_rows = {
        level: add_plus1_period_flag_to_rows(
            execution["level_results"][level]["rows"],
            current_start,
        )
        for level in CSV_LEVELS
    }
    comprehensive_rows = add_plus1_period_flag_to_rows(
        execution["comprehensive_results"],
        current_start,
    )
    filtered_comprehensive_rows = build_filtered_comprehensive_rows(
        comprehensive_rows,
        current_start,
    )
    filtered_comprehensive_csv_path = output_dir / FILTERED_COMPREHENSIVE_CSV
    write_level_csvs(output_dir, level_rows, comprehensive_rows, filtered_comprehensive_rows)
    notice_summary_source_rows = level_rows["notice"]
    filtered_notice_summary_source_rows = build_filtered_comprehensive_rows(
        notice_summary_source_rows,
        current_start,
    )
    summary_rows = build_label_poc_summary_rows(notice_summary_source_rows)
    filtered_summary_rows = build_label_poc_summary_rows(
        filtered_notice_summary_source_rows
    )
    summary_csv_path = output_dir / "汇总统计.csv"
    filtered_summary_csv_path = output_dir / FILTERED_SUMMARY_CSV
    write_summary_csv(summary_csv_path, summary_rows)
    write_summary_csv(filtered_summary_csv_path, filtered_summary_rows)
    workbook_path = write_workbook(
        output_dir,
        period,
        level_rows,
        comprehensive_rows,
        summary_rows,
        filtered_summary_rows,
        filtered_comprehensive_rows,
    )
    return ReportArtifacts(
        summary_rows=summary_rows,
        filtered_summary_rows=filtered_summary_rows,
        workbook_path=workbook_path,
        summary_csv_path=summary_csv_path,
        filtered_summary_csv_path=filtered_summary_csv_path,
        filtered_comprehensive_csv_path=filtered_comprehensive_csv_path,
        filtered_comprehensive_row_count=len(filtered_comprehensive_rows),
    )


def write_level_csvs(
    output_dir: Path,
    level_rows: dict[str, list[dict[str, Any]]],
    comprehensive_rows: list[dict[str, Any]],
    filtered_comprehensive_rows: list[dict[str, Any]],
) -> None:
    for level in CSV_LEVELS:
        write_level_csv(output_dir / f"{level}.csv", level_rows[level])
    write_level_csv(output_dir / "综合.csv", comprehensive_rows)
    write_level_csv(output_dir / FILTERED_COMPREHENSIVE_CSV, filtered_comprehensive_rows)


def write_level_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file, fieldnames=[header for _, header in LEVEL_COLUMN_SPECS]
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    header: csv_value(csv_column_value(row, field))
                    for field, header in LEVEL_COLUMN_SPECS
                }
            )


def write_summary_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file, fieldnames=[header for _, header in SUMMARY_COLUMN_SPECS]
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    header: csv_value(row.get(field))
                    for field, header in SUMMARY_COLUMN_SPECS
                }
            )


def write_workbook(
    output_dir: Path,
    period: dict[str, str],
    level_rows: dict[str, list[dict[str, Any]]],
    comprehensive_rows: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    filtered_summary_rows: list[dict[str, Any]],
    filtered_comprehensive_rows: list[dict[str, Any]],
) -> Path:
    workbook_path = (
        output_dir
        / f"low_label_rate_grading_{period['current_start']}_{period['current_end']}.xlsx"
    )
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, rows in [
        ("P0", level_rows["P0"]),
        ("P1", level_rows["P1"]),
        ("P2", level_rows["P2"]),
        ("Notice", level_rows["notice"]),
        ("综合", comprehensive_rows),
        ("综合_剔除+1同意", filtered_comprehensive_rows),
    ]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([header for _, header in LEVEL_COLUMN_SPECS])
        for row in rows:
            sheet.append(
                [
                    csv_value(csv_column_value(row, field))
                    for field, _ in LEVEL_COLUMN_SPECS
                ]
            )
        style_sheet_header(sheet)

    summary_sheet = workbook.create_sheet("汇总统计")
    summary_sheet.append([header for _, header in SUMMARY_COLUMN_SPECS])
    for row in summary_rows:
        summary_sheet.append(
            [csv_value(row.get(field)) for field, _ in SUMMARY_COLUMN_SPECS]
        )
    style_sheet_header(summary_sheet)

    filtered_summary_sheet = workbook.create_sheet("汇总统计_剔除+1同意")
    filtered_summary_sheet.append([header for _, header in SUMMARY_COLUMN_SPECS])
    for row in filtered_summary_rows:
        filtered_summary_sheet.append(
            [csv_value(row.get(field)) for field, _ in SUMMARY_COLUMN_SPECS]
        )
    style_sheet_header(filtered_summary_sheet)
    workbook.save(workbook_path)
    return workbook_path


def build_filtered_comprehensive_rows(
    rows: list[dict[str, Any]],
    current_start: str,
) -> list[dict[str, Any]]:
    return [
        row
        for row in rows
        if not is_pre_period_plus1_agreed(row, current_start)
    ]


def add_plus1_period_flag_to_rows(
    rows: list[dict[str, Any]],
    current_start: str,
) -> list[dict[str, Any]]:
    enriched_rows: list[dict[str, Any]] = []
    for row in rows:
        enriched = dict(row)
        enriched["plus1_agreed_before_current_period"] = (
            "是" if is_pre_period_plus1_agreed(row, current_start) else "否"
        )
        enriched_rows.append(enriched)
    return enriched_rows


def is_pre_period_plus1_agreed(row: dict[str, Any], current_start: str) -> bool:
    update_date = normalize_date_str(row.get("plus1_update_date"))
    return (
        row.get("is_plus1_agreed") == "是"
        and bool(update_date)
        and update_date < current_start
    )


def normalize_date_str(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text[:10].replace("/", "-")).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"Invalid ISO date value: {value!r}") from exc


def write_poc_routing_artifact(
    *,
    output_dir: Path,
    sample: dict[str, Any],
    source_path: Path,
    online_write_executed: bool,
) -> Path:
    poc_routing_path = output_dir / "poc_routing_plan.json"
    plan = build_poc_routing_plan(
        sample,
        source_stage_1_result=relative_to_root(source_path),
    )
    plan["routing_constraints"]["online_write_executed"] = online_write_executed
    write_json(poc_routing_path, plan)
    return poc_routing_path


def build_summary(
    *,
    source_path: Path,
    output_dir: Path,
    workbook_path: Path,
    sample: dict[str, Any],
    period: dict[str, str],
    sheet_url: str | None,
    top_n: int,
    self_send_requested: bool,
    label_poc_summary_count: int,
    filtered_label_poc_summary_count: int,
    filtered_comprehensive_row_count: int,
    online_write_attempted: bool,
    online_write_executed: bool,
) -> dict[str, Any]:
    execution = sample["readonly_execution"]
    provenance = sample["provenance"]
    return {
        "schema_version": "stage_2_notification_draft.v1",
        "report_type": REPORT_TYPE,
        "scenario_key": SCENARIO_KEY,
        "run_mode": "debug_only_with_explicit_self_send"
        if self_send_requested
        else "draft_only_with_name_level_poc_routing",
        "source_stage_1_result": relative_to_root(source_path),
        "reference_docs": [SCENARIO_REFERENCE],
        "asset_refs": {
            "poc_mapping": POC_MAPPING_ASSET,
            "card_template": CARD_TEMPLATE_ASSET,
            "card_schema_notes": CARD_SCHEMA_NOTES_ASSET,
        },
        "output_dir": relative_to_root(output_dir),
        "dataset_id": provenance["dataset_id"],
        "region": provenance["region"],
        "period": period,
        "level_counts": execution["level_counts"],
        "comprehensive_alert_count": execution["row_count"],
        "comprehensive_reason_count": execution["row_count"],
        "comprehensive_strategy_group_count": execution["row_count"],
        "comprehensive_exclude_pre_period_plus1_count": filtered_comprehensive_row_count,
        "plus1_exclusion_cutoff_date": period["current_start"],
        "fallback_reason": sample["QueryPlan"]["fallback_reason"],
        "metric_formula": execution["metric_formula"],
        "top_n": top_n,
        "sheet_url": sheet_url,
        "online_write_attempted": online_write_attempted,
        "online_write_executed": online_write_executed,
        "label_poc_summary_count": label_poc_summary_count,
        "label_poc_summary_exclude_pre_period_plus1_count": (
            filtered_label_poc_summary_count
        ),
        "outputs": {
            "summary_json": "summary.json",
            "notice_csv": "notice.csv",
            "P2_csv": "P2.csv",
            "P1_csv": "P1.csv",
            "P0_csv": "P0.csv",
            "comprehensive_csv": "综合.csv",
            "comprehensive_exclude_pre_period_plus1_csv": FILTERED_COMPREHENSIVE_CSV,
            "workbook": workbook_path.name,
            "poc_routing_plan": "poc_routing_plan.json",
            "notification_draft": "notification_draft.json",
            "send_plan": "send_plan.json",
            "summary_by_label_poc_csv": "汇总统计.csv",
            "summary_by_label_poc_exclude_pre_period_plus1_csv": FILTERED_SUMMARY_CSV,
        },
        "source_footer": sample["source_footer"],
    }


def write_card_artifacts(
    *,
    publish_dir: Path,
    summary: dict[str, Any],
    summary_rows: list[dict[str, Any]],
    level_results: dict[str, dict[str, Any]],
    top_n: int,
    sheet_url: str | None,
    title: str,
    full_evidence_rows: list[dict[str, Any]],
) -> CardArtifacts:
    template_contract = load_card_template_contract()
    level_top_rows = build_level_top_rows(level_results, top_n)
    summary_card_rows = build_card_summary_rows(summary_rows)
    hash_rows = [
        {
            "scenario_key": summary.get("scenario_key"),
            "report_type": summary.get("report_type"),
            "period": summary.get("period"),
            "query_plan_id": summary.get("source_footer", {}).get("query_plan_id"),
            "source_footer": summary.get("source_footer"),
            "sheet_url": sheet_url,
            "summary_rows": summary_card_rows,
            "full_evidence_rows": full_evidence_rows,
        }
    ]
    card_with_meta = render_grading_card(
        summary=summary,
        summary_rows=summary_card_rows,
        level_top_rows=level_top_rows,
        sheet_url=sheet_url,
        title=title,
        hash_input=hash_rows,
        template_contract=template_contract,
    )
    verify_card_hash(card_with_meta, hash_rows)
    card_json = strip_internal_keys(card_with_meta)
    hash_check = {
        "ok": True,
        "data_hash": card_with_meta["_meta"]["_data_hash"],
        "top_rows_count": sum(len(rows) for rows in level_top_rows.values()),
        "full_evidence_row_count": len(full_evidence_rows),
        "level_top_rows_count": {
            level: len(rows) for level, rows in level_top_rows.items()
        },
        "internal_meta_removed": "_meta" not in card_json,
        "design_check": card_design_check(card_with_meta),
    }

    card_path = publish_dir / f"{REPORT_TYPE}.card.json"
    card_with_meta_path = publish_dir / f"{REPORT_TYPE}.card.with_meta.json"
    hash_check_path = publish_dir / "card_hash_check.json"
    write_json(card_path, card_json, compact=True)
    write_json(card_with_meta_path, card_with_meta)
    write_json(hash_check_path, hash_check)
    return CardArtifacts(
        card_path=card_path,
        card_with_meta_path=card_with_meta_path,
        hash_check_path=hash_check_path,
        card_json=card_json,
        card_with_meta=card_with_meta,
        hash_check=hash_check,
    )


def load_card_template_contract() -> dict[str, Any]:
    path = (
        Path(__file__).resolve().parents[1]
        / "assets"
        / "efficiency-label-rate"
        / "low_efficiency_grading_card_template.json"
    )
    payload = load_json(path)
    if payload.get("schema") != "2.0":
        raise ValueError("Card template schema must be 2.0.")
    if payload.get("scenario_key") != SCENARIO_KEY:
        raise ValueError("Card template scenario_key mismatch.")
    return payload


def build_publish_summary(
    *,
    source_path: Path,
    output_dir: Path,
    summary_path: Path,
    workbook_path: Path,
    summary_csv_path: Path,
    filtered_summary_csv_path: Path,
    sheet_url: str | None,
    card: CardArtifacts,
    notification_draft_path: Path,
    send_plan_path: Path,
    identity: str,
    sent_payload: dict[str, Any] | None,
    target_user_id: str | None,
    target_chat_id: str | None,
    online_write_attempted: bool,
    online_write_executed: bool,
) -> dict[str, Any]:
    return {
        "report_type": REPORT_TYPE,
        "scenario_key": SCENARIO_KEY,
        "source_stage_1_result": relative_to_root(source_path),
        "output_dir": relative_to_root(output_dir),
        "summary_json": relative_to_root(summary_path),
        "workbook": relative_to_root(workbook_path),
        "summary_by_label_poc_csv": relative_to_root(summary_csv_path),
        "summary_by_label_poc_exclude_pre_period_plus1_csv": relative_to_root(
            filtered_summary_csv_path
        ),
        "sheet_url": sheet_url,
        "online_write_attempted": online_write_attempted,
        "online_write_executed": online_write_executed,
        "card_json": relative_to_root(card.card_path),
        "card_json_with_meta": relative_to_root(card.card_with_meta_path),
        "card_hash_check": relative_to_root(card.hash_check_path),
        "sent": sent_payload is not None,
        "send_identity": identity if sent_payload is not None else None,
        "target_type": "user" if target_user_id else ("chat" if target_chat_id else None),
        "target_user": "self" if target_user_id and sent_payload is not None else None,
        "target_user_open_id_prefix": mask_identifier(target_user_id)
        if target_user_id and sent_payload is not None
        else None,
        "target_chat_id": target_chat_id if sent_payload is not None else None,
        "message_id": extract_message_id(sent_payload),
        "send_result": sanitize_send_payload(sent_payload),
        "notification_draft": relative_to_root(notification_draft_path),
        "send_plan": relative_to_root(send_plan_path),
    }


def build_notification_draft(
    *,
    summary: dict[str, Any],
    poc_routing_path: Path,
    card_path: Path,
    card_with_meta_path: Path,
    hash_check_path: Path,
    publish_summary: dict[str, Any],
) -> dict[str, Any]:
    if not poc_routing_path.exists():
        raise FileNotFoundError(f"Missing POC routing plan: {poc_routing_path}")
    poc_routing = load_json(poc_routing_path)
    constraints = poc_routing.get("routing_constraints", {})
    return {
        "schema_version": "stage_2_notification_draft_detail.v1",
        "scenario_key": SCENARIO_KEY,
        "report_type": REPORT_TYPE,
        "draft_mode": "self_preview_with_name_level_poc_routing",
        "default_self_validation": True,
        "real_poc_mapping_used": poc_routing.get("real_poc_mapping_used", False),
        "source_stage_1_result": summary["source_stage_1_result"],
        "level_counts": summary["level_counts"],
        "comprehensive_reason_count": summary["comprehensive_reason_count"],
        "comprehensive_alert_count": summary["comprehensive_alert_count"],
        "comprehensive_strategy_group_count": summary[
            "comprehensive_strategy_group_count"
        ],
        "data_link": {
            "sheet_url": summary.get("sheet_url"),
            "workbook": summary["outputs"].get("workbook"),
            "csv_files": {
                "summary_by_label_poc": summary["outputs"].get(
                    "summary_by_label_poc_csv"
                ),
                "summary_by_label_poc_exclude_pre_period_plus1": summary[
                    "outputs"
                ].get("summary_by_label_poc_exclude_pre_period_plus1_csv"),
                "notice": summary["outputs"].get("notice_csv"),
                "P2": summary["outputs"].get("P2_csv"),
                "P1": summary["outputs"].get("P1_csv"),
                "P0": summary["outputs"].get("P0_csv"),
                "comprehensive": summary["outputs"].get("comprehensive_csv"),
                "comprehensive_exclude_pre_period_plus1": summary["outputs"].get(
                    "comprehensive_exclude_pre_period_plus1_csv"
                ),
            },
        },
        "card_draft": {
            "card_json": relative_to_root(card_path),
            "card_json_with_meta": relative_to_root(card_with_meta_path),
            "card_hash_check": relative_to_root(hash_check_path),
            "send_card_meta_removed": True,
        },
        "poc_routing": {
            "poc_routing_plan": relative_to_root(poc_routing_path),
            "routing_mode": poc_routing.get("routing_mode"),
            "fallback_to_default_user": poc_routing.get("fallback_to_default_user"),
            "default_recipient": poc_routing.get("default_recipient"),
            "routing_key": poc_routing.get("routing_key"),
            "contact_resolution_status": poc_routing.get("contact_resolution_status"),
            "requires_contact_resolution_before_real_send": poc_routing.get(
                "requires_contact_resolution_before_real_send", True
            ),
            "mapped_row_count": poc_routing.get("mapped_row_count"),
            "unmapped_row_count": poc_routing.get("unmapped_row_count"),
            "missing_route_dimension_count": poc_routing.get(
                "missing_route_dimension_count"
            ),
            "poc_summary": poc_routing.get("poc_summary", []),
            "routing_rules": summarize_routing_rules(poc_routing),
            "routing_constraints": constraints,
        },
        "methodology": {
            "metric_formula": summary.get("metric_formula"),
            "period": summary.get("period"),
            "source_footer": summary.get("source_footer"),
            "reference_docs": [SCENARIO_REFERENCE],
            "asset_refs": {
                "poc_mapping": POC_MAPPING_ASSET,
                "card_template": CARD_TEMPLATE_ASSET,
                "card_schema_notes": CARD_SCHEMA_NOTES_ASSET,
            },
        },
        "send_safety": {
            "current_stage": "draft_only_for_group_send",
            "current_preview_recipient": "self",
            "requires_confirmation_before_group_send": True,
            "group_send_blocked": constraints.get("group_send_blocked", True),
            "group_send_allowed": constraints.get("group_send_allowed", False),
            "sent": False,
            "real_group_send_executed": False,
            "online_write_executed": constraints.get("online_write_executed", False),
            "self_preview_sent": publish_summary.get("sent", False),
            "self_preview_message_id": publish_summary.get("message_id"),
        },
        "escalation_draft": {
            "required_levels": [
                level
                for level in ("P0", "P1")
                if int(summary.get("level_counts", {}).get(level, 0)) > 0
            ],
            "message": "请相关负责人复核低效策略证据、处理计划和完成时间。",
            "send_allowed": False,
        },
        "evidence_refs": {
            "analysis_artifact": summary["source_stage_1_result"],
            "query_plan_id": summary.get("source_footer", {}).get("query_plan_id"),
            "source_footer": summary.get("source_footer"),
            "card_hash_check": relative_to_root(hash_check_path),
            "reports": summary.get("outputs", {}),
        },
        "failure_branches": {
            "unmapped_poc": "fallback_to_self_preview_and_require_confirmation",
            "missing_open_id": "block_real_send_until_contact_resolution",
            "missing_sheet_url": "keep_local_workbook_and_csv_links",
            "real_group_send_requested": "keep_group_send_blocked",
            "card_hash_mismatch": "regenerate_card_before_send",
        },
        "provenance": {
            "reference_docs": [SCENARIO_REFERENCE],
            "asset_refs": {
                "poc_mapping": POC_MAPPING_ASSET,
                "card_template": CARD_TEMPLATE_ASSET,
                "card_schema_notes": CARD_SCHEMA_NOTES_ASSET,
            },
            "dataset_id": summary.get("dataset_id"),
            "region": summary.get("region"),
            "query_plan_id": summary.get("source_footer", {}).get("query_plan_id"),
            "online_write_attempted": summary.get("online_write_attempted", False),
            "online_write_executed": summary.get("online_write_executed", False),
        },
    }


def build_send_plan(
    *,
    identity: str,
    publish_summary: dict[str, Any],
    poc_routing_path: Path,
    card_path: Path,
    notification_draft_path: Path,
) -> dict[str, Any]:
    return {
        "schema_version": "stage_2_send_plan.v1",
        "scenario_key": SCENARIO_KEY,
        "report_type": REPORT_TYPE,
        "plan_mode": "manual_confirmation_required",
        "target_type": "future_group_or_name_level_poc_after_confirmation",
        "target_source": relative_to_root(poc_routing_path),
        "content_source": {
            "card_json": relative_to_root(card_path),
            "notification_draft": relative_to_root(notification_draft_path),
        },
        "reference_docs": [SCENARIO_REFERENCE],
        "asset_refs": {
            "poc_mapping": POC_MAPPING_ASSET,
            "card_template": CARD_TEMPLATE_ASSET,
            "card_schema_notes": CARD_SCHEMA_NOTES_ASSET,
        },
        "send_identity": identity,
        "requires_confirmation": True,
        "confirmation_status": "not_requested",
        "group_send_blocked": True,
        "group_send_allowed": False,
        "group_recipients": [],
        "sent": False,
        "real_group_send_executed": False,
        "online_write_attempted": publish_summary.get(
            "online_write_attempted", False
        ),
        "online_write_executed": publish_summary.get(
            "online_write_executed", False
        ),
        "blocked_reason": (
            "The notification artifact has name-level POC routing only. Feishu open_id resolution, "
            "target chat confirmation, and explicit confirmation are required before "
            "real POC/group sending."
        ),
        "self_preview": {
            "default_recipient": "self",
            "sent": publish_summary.get("sent", False),
            "message_id": publish_summary.get("message_id"),
        },
        "required_before_real_send": [
            "resolve POC names to confirmed Feishu open_id",
            "confirm target chat or POC recipients",
            "confirm send identity and content",
            "run validator with group-send gate enabled",
        ],
    }


def build_label_poc_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            str(row.get("data_source") or ""),
            str(row.get("mach_root_label_name", "")),
            str(row.get("POC") or row.get("poc_name") or "未映射"),
        )
        bucket = grouped.setdefault(
            key,
            {
                "data_source": key[0],
                "mach_root_label_name": key[1],
                "POC": key[2],
                "_strategy_names": set(),
                "_total_review_in_cnt": 0.0,
                "_total_review_done_cnt": 0.0,
                "_total_label_cnt": 0.0,
                "_avg_review_in_cnt": 0.0,
                "_avg_review_done_cnt": 0.0,
                "_avg_label_cnt": 0.0,
            },
        )
        strategy_name = str(row.get("strategy_name") or "")
        strategy_id = str(row.get("strategy_id") or "")
        strategy_key = strategy_id or strategy_name
        if strategy_key:
            bucket["_strategy_names"].add(strategy_key)
        bucket["_total_review_in_cnt"] += float(row.get("total_review_in_cnt", 0) or 0)
        bucket["_total_review_done_cnt"] += float(
            row.get("total_review_done_cnt", 0) or 0
        )
        bucket["_total_label_cnt"] += float(row.get("total_label_cnt", 0) or 0)
        bucket["_avg_review_in_cnt"] += float(row.get("avg_review_in_cnt", 0) or 0)
        bucket["_avg_review_done_cnt"] += float(row.get("avg_review_done_cnt", 0) or 0)
        bucket["_avg_label_cnt"] += float(row.get("avg_label_cnt", 0) or 0)

    result: list[dict[str, Any]] = []
    for bucket in grouped.values():
        avg_review_in_cnt = round(bucket["_avg_review_in_cnt"])
        avg_review_done_cnt = round(bucket["_avg_review_done_cnt"])
        avg_label_cnt = round(bucket["_avg_label_cnt"])
        result.append(
            {
                "data_source": bucket["data_source"],
                "mach_root_label_name": bucket["mach_root_label_name"],
                "POC": bucket["POC"],
                "low_efficiency_strategy_count": len(bucket["_strategy_names"]),
                "avg_review_in_cnt": avg_review_in_cnt,
                "avg_review_done_cnt": avg_review_done_cnt,
                "avg_label_cnt": avg_label_cnt,
                "label_rate": avg_label_cnt / avg_review_done_cnt
                if avg_review_done_cnt
                else 0.0,
            }
        )
    return sorted(
        result,
        key=lambda item: (
            -float(item["avg_review_done_cnt"]),
            str(item["data_source"]),
            str(item["mach_root_label_name"]),
            str(item["POC"]),
        ),
    )


def build_level_top_rows(
    level_results: dict[str, dict[str, Any]],
    top_n: int,
) -> dict[str, list[dict[str, Any]]]:
    return {
        level: build_top_rows(level_results.get(level, {}).get("rows", []), top_n)
        for level in CARD_LEVELS
    }


def flatten_level_top_rows(
    level_top_rows: dict[str, list[dict[str, Any]]]
) -> list[dict[str, Any]]:
    flattened: list[dict[str, Any]] = []
    for level in CARD_LEVELS:
        flattened.extend(level_top_rows.get(level, []))
    return flattened


def build_card_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "data_source": row.get("data_source", ""),
            "mach_root_label_name": row.get("mach_root_label_name", ""),
            "POC": row.get("POC", ""),
            "low_efficiency_strategy_count": int(
                row.get("low_efficiency_strategy_count", 0) or 0
            ),
            "avg_review_in_cnt": int(round(float(row.get("avg_review_in_cnt", 0) or 0))),
            "avg_review_done_cnt": int(
                round(float(row.get("avg_review_done_cnt", 0) or 0))
            ),
            "avg_label_cnt": int(round(float(row.get("avg_label_cnt", 0) or 0))),
            "label_rate": pct_value(row.get("label_rate")),
        }
        for row in rows
    ]


def build_top_rows(rows: list[dict[str, Any]], top_n: int) -> list[dict[str, Any]]:
    top_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows[:top_n], 1):
        level = row["severity_level"]
        top_rows.append(
            {
                "rank": index,
                "level": [{"text": level, "color": LEVEL_COLORS.get(level, "blue")}],
                "poc_name": row.get("POC") or row.get("poc_name", "未映射"),
                "data_source": row.get("data_source", ""),
                "warning_dimension": row.get("warning_dimension", ""),
                "mach_root_label_name": row.get("mach_root_label_name", ""),
                "strategy_id": row.get("strategy_id", ""),
                "strategy_name": row.get("strategy_name", ""),
                "max_data_date": row.get("max_data_date", ""),
                "avg_in": round(float(row["avg_review_in_cnt"])),
                "avg_done": round(float(row["avg_review_done_cnt"])),
                "avg_labeled": round(float(row["avg_label_cnt"])),
                "label_rate": pct_value(row.get("label_rate")),
                "hit_reason": csv_value(row.get("hit_conditions")),
            }
        )
    return top_rows


def summarize_routing_rules(poc_routing: dict[str, Any]) -> dict[str, Any]:
    rules = poc_routing.get("routing_rules", {})
    return {
        level: {
            "target_roles": rule.get("target_roles", []),
            "action_required": rule.get("action_required"),
            "default_recipient": rule.get("default_recipient"),
            "recipient_resolution": rule.get("recipient_resolution", {}),
            "requires_human_confirmation_before_real_send": rule.get(
                "requires_human_confirmation_before_real_send"
            ),
            "group_send_blocked": rule.get("group_send_blocked"),
            "alert_count": rule.get("alert_count", rule.get("strategy_group_count")),
            "reason_count": rule.get("reason_count"),
            "strategy_group_count": rule.get("strategy_group_count"),
            "poc_names": rule.get("poc_names", []),
            "unmapped_labels": rule.get("unmapped_labels", []),
        }
        for level, rule in rules.items()
    }


def style_sheet_header(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="D9E8FF")
    font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"


def csv_value(value: Any) -> Any:
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    return value


def csv_column_value(row: dict[str, Any], column: str) -> Any:
    if column == "POC":
        return row.get("POC") or row.get("poc_name")
    return row.get(column)


def pct_value(value: Any) -> str:
    try:
        return f"{float(value) * 100:.2f}%"
    except (TypeError, ValueError):
        return str(value)


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = (
        json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        if compact
        else json.dumps(value, ensure_ascii=False, indent=2)
    )
    path.write_text(text + "\n", encoding="utf-8")


def relative_to_root(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return str(resolved)


def mask_identifier(value: str | None) -> str | None:
    if not value:
        return None
    return value[:10] + "..."


def extract_message_id(payload: dict[str, Any] | None) -> str | None:
    if not payload:
        return None
    data = payload.get("data", {})
    if isinstance(data, dict):
        return data.get("message_id")
    return None


def sanitize_send_payload(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not payload:
        return None
    data = payload.get("data", {})
    return {
        "ok": payload.get("ok"),
        "identity": payload.get("identity"),
        "message_id": data.get("message_id") if isinstance(data, dict) else None,
        "create_time": data.get("create_time") if isinstance(data, dict) else None,
    }


def safe_idempotency_key(raw: str) -> str:
    key = re.sub(r"[^A-Za-z0-9-]+", "-", raw).strip("-")
    key = re.sub(r"-{2,}", "-", key)
    return key[-50:] or "label-rate-card"


if __name__ == "__main__":
    main()
