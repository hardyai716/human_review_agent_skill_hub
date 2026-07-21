#!/usr/bin/env python3
"""Build a two-period filtered label-rate summary comparison workbook."""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sheet_importer import import_xlsx_as_feishu_sheet


FILTERED_SUMMARY_FILENAME = "汇总统计_剔除+1同意.csv"
COMPARISON_SHEET_NAME = "汇总统计_剔除+1同意对比"
COMPARISON_SUMMARY_FILENAME = "weekly_summary_comparison.json"
REQUIRED_COLUMNS = (
    "机审一级标签",
    "POC",
    "低效策略数",
    "低效策略日均进审量",
    "低效策略日均完审量",
    "低效策略日均打标量",
    "低效策略打标率",
)


@dataclass(frozen=True)
class SummaryRow:
    """A label and POC row from a filtered weekly summary."""

    mach_root_label_name: str
    poc: str
    low_efficiency_strategy_count: int
    avg_review_in_cnt: int
    avg_review_done_cnt: int
    avg_label_cnt: int

    @property
    def label_rate(self) -> float | None:
        if self.avg_review_done_cnt <= 0:
            return None
        return self.avg_label_cnt / self.avg_review_done_cnt


@dataclass(frozen=True)
class WeeklyComparisonArtifacts:
    """Files and structured facts produced by a weekly comparison."""

    output_dir: Path
    workbook_path: Path
    summary_path: Path
    comparison_rows: list[dict[str, Any]]
    totals: dict[str, Any]
    sheet_url: str | None
    online_write_attempted: bool
    online_write_executed: bool


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Build a screenshot-style comparison workbook from two "
            "汇总统计_剔除+1同意 CSV files. This script never sends messages."
        )
    )
    parser.add_argument("--previous-summary", required=True)
    parser.add_argument("--current-summary", required=True)
    parser.add_argument("--previous-start-date", required=True)
    parser.add_argument("--previous-end-date", required=True)
    parser.add_argument("--current-start-date", required=True)
    parser.add_argument("--current-end-date", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--previous-label")
    parser.add_argument("--current-label")
    parser.add_argument(
        "--import-sheet",
        action="store_true",
        help=(
            "Opt in to importing the generated XLSX as a Feishu online sheet. "
            "Off by default because it is an online write."
        ),
    )
    parser.add_argument("--sheet-name")
    args = parser.parse_args()

    artifacts = build_weekly_summary_comparison(
        previous_summary_path=Path(args.previous_summary),
        current_summary_path=Path(args.current_summary),
        previous_start_date=args.previous_start_date,
        previous_end_date=args.previous_end_date,
        current_start_date=args.current_start_date,
        current_end_date=args.current_end_date,
        output_dir=Path(args.output_dir),
        previous_label=args.previous_label,
        current_label=args.current_label,
        auto_import_sheet=args.import_sheet,
        sheet_name=args.sheet_name,
    )
    print(
        json.dumps(
            {
                "schema_version": "label_rate_weekly_summary_comparison.v1",
                "workbook": str(artifacts.workbook_path),
                "summary": str(artifacts.summary_path),
                "sheet_url": artifacts.sheet_url,
                "online_write_attempted": artifacts.online_write_attempted,
                "online_write_executed": artifacts.online_write_executed,
                "totals": artifacts.totals,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def build_weekly_summary_comparison(
    *,
    previous_summary_path: Path,
    current_summary_path: Path,
    previous_start_date: str,
    previous_end_date: str,
    current_start_date: str,
    current_end_date: str,
    output_dir: Path,
    previous_label: str | None = None,
    current_label: str | None = None,
    auto_import_sheet: bool = False,
    sheet_name: str | None = None,
) -> WeeklyComparisonArtifacts:
    """Build a visual weekly comparison from two filtered summary snapshots."""

    previous_period = normalize_period(previous_start_date, previous_end_date)
    current_period = normalize_period(current_start_date, current_end_date)
    previous_rows = read_filtered_summary(previous_summary_path)
    current_rows = read_filtered_summary(current_summary_path)
    previous_index = index_rows(previous_rows)
    current_index = index_rows(current_rows)
    ordered_keys = ordered_comparison_keys(
        previous_rows=previous_rows,
        current_rows=current_rows,
    )

    comparison_rows = build_comparison_rows(
        ordered_keys=ordered_keys,
        previous_index=previous_index,
        current_index=current_index,
    )
    totals = build_totals(previous_rows, current_rows)

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / comparison_workbook_filename(
        previous_period=previous_period,
        current_period=current_period,
    )
    write_comparison_workbook(
        workbook_path=workbook_path,
        comparison_rows=comparison_rows,
        totals=totals,
        previous_period=previous_period,
        current_period=current_period,
        previous_label=previous_label or display_period(previous_period),
        current_label=current_label or display_period(current_period),
    )

    online_write_attempted = bool(auto_import_sheet)
    sheet_url: str | None = None
    if auto_import_sheet:
        sheet_url = import_xlsx_as_feishu_sheet(
            workbook_path=workbook_path,
            output_dir=output_dir,
            sheet_name=sheet_name
            or (
                "低效打标汇总统计_剔除+1同意_"
                f"{previous_period['start']}_{previous_period['end']}_vs_"
                f"{current_period['start']}_{current_period['end']}"
            ),
            result_filename="comparison_sheet_import_result.json",
        )

    summary_path = output_dir / COMPARISON_SUMMARY_FILENAME
    payload = {
        "schema_version": "label_rate_weekly_summary_comparison.v1",
        "report_type": "filtered_summary_weekly_comparison",
        "summary_source_filename": FILTERED_SUMMARY_FILENAME,
        "periods": {
            "previous": {
                **previous_period,
                "label": previous_label or display_period(previous_period),
                "source_csv": str(previous_summary_path),
            },
            "current": {
                **current_period,
                "label": current_label or display_period(current_period),
                "source_csv": str(current_summary_path),
            },
        },
        "comparison_rows": comparison_rows,
        "totals": totals,
        "workbook": workbook_path.name,
        "sheet_url": sheet_url,
        "online_write_attempted": online_write_attempted,
        "online_write_executed": bool(sheet_url),
        "source_footer": (
            "数据源：各周期真实只读全等级分级结果的"
            "「汇总统计_剔除+1同意」。每周期独立剔除更新日期早于"
            "本周期开始日的 +1 同意策略；低效策略打标率按"
            "日均打标量 / 日均完审量加权计算。"
        ),
    }
    write_json(summary_path, payload)
    return WeeklyComparisonArtifacts(
        output_dir=output_dir,
        workbook_path=workbook_path,
        summary_path=summary_path,
        comparison_rows=comparison_rows,
        totals=totals,
        sheet_url=sheet_url,
        online_write_attempted=online_write_attempted,
        online_write_executed=bool(sheet_url),
    )


def normalize_period(start_date: str, end_date: str) -> dict[str, str]:
    start = date.fromisoformat(start_date.strip().replace("/", "-"))
    end = date.fromisoformat(end_date.strip().replace("/", "-"))
    if end < start:
        raise ValueError("period end date must not be earlier than start date.")
    return {"start": start.isoformat(), "end": end.isoformat()}


def display_period(period: dict[str, str]) -> str:
    start = date.fromisoformat(period["start"])
    end = date.fromisoformat(period["end"])
    return f"{start.month}.{start.day}-{end.month}.{end.day}"


def comparison_workbook_filename(
    *,
    previous_period: dict[str, str],
    current_period: dict[str, str],
) -> str:
    return (
        "低效打标汇总统计_剔除+1同意_"
        f"{previous_period['start'].replace('-', '')}_"
        f"{previous_period['end'].replace('-', '')}_vs_"
        f"{current_period['start'].replace('-', '')}_"
        f"{current_period['end'].replace('-', '')}.xlsx"
    )


def read_filtered_summary(path: Path) -> list[SummaryRow]:
    if path.name != FILTERED_SUMMARY_FILENAME:
        raise ValueError(
            f"expected {FILTERED_SUMMARY_FILENAME}, got {path.name!r}."
        )
    with path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        fieldnames = set(reader.fieldnames or [])
        missing = [column for column in REQUIRED_COLUMNS if column not in fieldnames]
        if missing:
            raise ValueError(f"{path} is missing required columns: {missing}")
        rows = [
            SummaryRow(
                mach_root_label_name=require_text(
                    raw_row.get("机审一级标签"),
                    path=path,
                    field="机审一级标签",
                ),
                poc=require_text(raw_row.get("POC"), path=path, field="POC"),
                low_efficiency_strategy_count=parse_non_negative_int(
                    raw_row.get("低效策略数"),
                    path=path,
                    field="低效策略数",
                ),
                avg_review_in_cnt=parse_non_negative_int(
                    raw_row.get("低效策略日均进审量"),
                    path=path,
                    field="低效策略日均进审量",
                ),
                avg_review_done_cnt=parse_non_negative_int(
                    raw_row.get("低效策略日均完审量"),
                    path=path,
                    field="低效策略日均完审量",
                ),
                avg_label_cnt=parse_non_negative_int(
                    raw_row.get("低效策略日均打标量"),
                    path=path,
                    field="低效策略日均打标量",
                ),
            )
            for raw_row in reader
        ]
    if not rows:
        raise ValueError(f"{path} has no summary rows.")
    return rows


def require_text(value: Any, *, path: Path, field: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{path} contains an empty {field}.")
    return text


def parse_non_negative_int(value: Any, *, path: Path, field: str) -> int:
    try:
        parsed = int(round(float(value or 0)))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{path} has invalid {field}: {value!r}") from exc
    if parsed < 0:
        raise ValueError(f"{path} has negative {field}: {value!r}")
    return parsed


def index_rows(rows: list[SummaryRow]) -> dict[tuple[str, str], SummaryRow]:
    index: dict[tuple[str, str], SummaryRow] = {}
    for row in rows:
        key = (row.mach_root_label_name, row.poc)
        if key in index:
            raise ValueError(f"duplicate summary key: {key}")
        index[key] = row
    return index


def ordered_comparison_keys(
    *,
    previous_rows: list[SummaryRow],
    current_rows: list[SummaryRow],
) -> list[tuple[str, str]]:
    previous_keys = [
        (row.mach_root_label_name, row.poc)
        for row in previous_rows
    ]
    previous_key_set = set(previous_keys)
    current_only = sorted(
        (
            row
            for row in current_rows
            if (row.mach_root_label_name, row.poc) not in previous_key_set
        ),
        key=lambda row: (
            -row.avg_review_done_cnt,
            row.mach_root_label_name,
            row.poc,
        ),
    )
    return [
        *previous_keys,
        *((row.mach_root_label_name, row.poc) for row in current_only),
    ]


def build_comparison_rows(
    *,
    ordered_keys: list[tuple[str, str]],
    previous_index: dict[tuple[str, str], SummaryRow],
    current_index: dict[tuple[str, str], SummaryRow],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key in ordered_keys:
        previous = previous_index.get(key)
        current = current_index.get(key)
        previous_done = previous.avg_review_done_cnt if previous else 0
        current_done = current.avg_review_done_cnt if current else 0
        done_delta = current_done - previous_done
        rows.append(
            {
                "mach_root_label_name": key[0],
                "poc": key[1],
                "previous_strategy_count": (
                    previous.low_efficiency_strategy_count if previous else 0
                ),
                "current_strategy_count": (
                    current.low_efficiency_strategy_count if current else 0
                ),
                "previous_avg_review_done_cnt": previous_done,
                "current_avg_review_done_cnt": current_done,
                "avg_review_done_delta": done_delta,
                "avg_review_done_growth_rate": (
                    done_delta / previous_done if previous_done else None
                ),
                "previous_label_rate": previous.label_rate if previous else None,
                "current_label_rate": current.label_rate if current else None,
            }
        )
    return rows


def build_totals(
    previous_rows: list[SummaryRow],
    current_rows: list[SummaryRow],
) -> dict[str, Any]:
    previous_strategy_count = sum(
        row.low_efficiency_strategy_count for row in previous_rows
    )
    current_strategy_count = sum(
        row.low_efficiency_strategy_count for row in current_rows
    )
    previous_done = sum(row.avg_review_done_cnt for row in previous_rows)
    current_done = sum(row.avg_review_done_cnt for row in current_rows)
    previous_label = sum(row.avg_label_cnt for row in previous_rows)
    current_label = sum(row.avg_label_cnt for row in current_rows)
    done_delta = current_done - previous_done
    return {
        "previous_strategy_count": previous_strategy_count,
        "current_strategy_count": current_strategy_count,
        "previous_avg_review_done_cnt": previous_done,
        "current_avg_review_done_cnt": current_done,
        "avg_review_done_delta": done_delta,
        "avg_review_done_growth_rate": (
            done_delta / previous_done if previous_done else None
        ),
        "previous_label_rate": previous_label / previous_done if previous_done else None,
        "current_label_rate": current_label / current_done if current_done else None,
    }


def write_comparison_workbook(
    *,
    workbook_path: Path,
    comparison_rows: list[dict[str, Any]],
    totals: dict[str, Any],
    previous_period: dict[str, str],
    current_period: dict[str, str],
    previous_label: str,
    current_label: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = COMPARISON_SHEET_NAME
    sheet.freeze_panes = "C3"
    sheet.sheet_view.showGridLines = False

    apply_grouped_headers(
        sheet=sheet,
        previous_label=previous_label,
        current_label=current_label,
    )
    for row_index, row in enumerate(comparison_rows, start=3):
        write_comparison_row(sheet=sheet, row_index=row_index, row=row)
    total_row_index = 3 + len(comparison_rows)
    write_total_row(sheet=sheet, row_index=total_row_index, totals=totals)
    write_source_footer(
        sheet=sheet,
        row_index=total_row_index + 2,
        previous_period=previous_period,
        current_period=current_period,
    )
    set_column_widths(sheet)
    workbook.save(workbook_path)


def apply_grouped_headers(
    *,
    sheet: Any,
    previous_label: str,
    current_label: str,
) -> None:
    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:D1")
    sheet.merge_cells("E1:H1")
    sheet.merge_cells("I1:J1")
    sheet["A1"] = "机审一级标签"
    sheet["B1"] = "POC"
    sheet["C1"] = "低效策略数"
    sheet["E1"] = "低效策略完审量"
    sheet["I1"] = "低效策略打标率"
    for column, value in enumerate(
        [
            previous_label,
            current_label,
            previous_label,
            current_label,
            "增量",
            "增幅",
            previous_label,
            current_label,
        ],
        start=3,
    ):
        sheet.cell(row=2, column=column, value=value)

    group_fills = {
        "dimension": PatternFill("solid", fgColor="D9E2F3"),
        "count": PatternFill("solid", fgColor="E2F0D9"),
        "volume": PatternFill("solid", fgColor="DDEBF7"),
        "rate": PatternFill("solid", fgColor="FCE4D6"),
    }
    for row_index in (1, 2):
        for column in range(1, 11):
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = (
                group_fills["dimension"]
                if column <= 2
                else group_fills["count"]
                if column <= 4
                else group_fills["volume"]
                if column <= 8
                else group_fills["rate"]
            )
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = header_border()
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 22


def write_comparison_row(*, sheet: Any, row_index: int, row: dict[str, Any]) -> None:
    values = [
        row["mach_root_label_name"],
        display_poc(str(row["poc"])),
        row["previous_strategy_count"],
        row["current_strategy_count"],
        row["previous_avg_review_done_cnt"],
        row["current_avg_review_done_cnt"],
        row["avg_review_done_delta"],
        row["avg_review_done_growth_rate"]
        if row["avg_review_done_growth_rate"] is not None
        else "/",
        row["previous_label_rate"] if row["previous_label_rate"] is not None else "/",
        row["current_label_rate"] if row["current_label_rate"] is not None else "/",
    ]
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column, value=value)
        cell.border = data_border()
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if column in {3, 4, 5, 6, 7}:
            cell.number_format = "#,##0"
        elif column in {8, 9, 10} and isinstance(value, float):
            cell.number_format = "0.0%"
        if column == 7 and int(row["avg_review_done_delta"]) > 0:
            cell.fill = PatternFill("solid", fgColor="F4CCCC")
            cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
    sheet.row_dimensions[row_index].height = 21


def write_total_row(*, sheet: Any, row_index: int, totals: dict[str, Any]) -> None:
    values = [
        "总计",
        "",
        totals["previous_strategy_count"],
        totals["current_strategy_count"],
        totals["previous_avg_review_done_cnt"],
        totals["current_avg_review_done_cnt"],
        totals["avg_review_done_delta"],
        totals["avg_review_done_growth_rate"]
        if totals["avg_review_done_growth_rate"] is not None
        else "/",
        totals["previous_label_rate"] if totals["previous_label_rate"] is not None else "/",
        totals["current_label_rate"] if totals["current_label_rate"] is not None else "/",
    ]
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.border = data_border()
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if column in {3, 4, 5, 6, 7}:
            cell.number_format = "#,##0"
        elif column in {8, 9, 10} and isinstance(value, float):
            cell.number_format = "0.0%"
        if column == 7 and int(totals["avg_review_done_delta"]) > 0:
            cell.fill = PatternFill("solid", fgColor="F4CCCC")
            cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
    sheet.row_dimensions[row_index].height = 22


def write_source_footer(
    *,
    sheet: Any,
    row_index: int,
    previous_period: dict[str, str],
    current_period: dict[str, str],
) -> None:
    sheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=10,
    )
    sheet.cell(
        row=row_index,
        column=1,
        value=(
            "数据源：各周期真实只读全等级分级结果的"
            "「汇总统计_剔除+1同意」。口径："
            f"{previous_period['start']}~{previous_period['end']} 与 "
            f"{current_period['start']}~{current_period['end']} 分别剔除"
            "更新日期早于本周期开始日的 +1 同意策略；"
            "打标率 = 日均打标量 / 日均完审量。"
        ),
    )
    cell = sheet.cell(row=row_index, column=1)
    cell.fill = PatternFill("solid", fgColor="F8F9FA")
    cell.border = data_border()
    cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    sheet.row_dimensions[row_index].height = 34


def display_poc(poc: str) -> str:
    return poc if poc == "未映射" else f"@{poc}"


def header_border() -> Border:
    side = Side(style="medium", color="5B9BD5")
    return Border(left=side, right=side, top=side, bottom=side)


def data_border() -> Border:
    side = Side(style="thin", color="333333")
    return Border(left=side, right=side, top=side, bottom=side)


def set_column_widths(sheet: Any) -> None:
    widths = {
        1: 26,
        2: 15,
        3: 13,
        4: 13,
        5: 16,
        6: 16,
        7: 13,
        8: 12,
        9: 14,
        10: 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def write_json(path: Path, value: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
