#!/usr/bin/env python3
"""Generate and optionally send a stage 2 label-rate grading card draft."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
NOTIFICATION_SCRIPTS = ROOT / "skills" / "notification" / "scripts"
sys.path.insert(0, str(NOTIFICATION_SCRIPTS))

from card_hash import strip_internal_keys, verify_card_hash  # noqa: E402
from render_label_rate_grading_card import (  # noqa: E402
    LEVEL_COLORS,
    card_design_check,
    render_grading_card,
)


SCENARIO_KEY = "efficiency-label-rate"
OUTPUT_DATE = "20260709"
DEFAULT_SOURCE = (
    ROOT
    / "evals"
    / SCENARIO_KEY
    / "stage_1_runs"
    / "20260708_real_readonly_label_rate_grading_results.jsonl"
)
DEFAULT_OUTPUT_DIR = (
    ROOT
    / "evals"
    / SCENARIO_KEY
    / "stage_2_runs"
    / f"{OUTPUT_DATE}_low_label_rate_grading_notification_draft"
)
REPORT_TYPE = "low_efficiency_grading"
LEVELS = ["P0", "P1", "P2", "notice"]
CSV_COLUMNS = [
    "severity_level",
    "severity_priority",
    "mach_root_label_name",
    "strategy_id",
    "strategy_name",
    "reason",
    "POC",
    "data_days",
    "total_review_in_cnt",
    "total_review_done_cnt",
    "total_label_cnt",
    "avg_review_in_cnt",
    "avg_review_done_cnt",
    "avg_label_cnt",
    "label_rate",
    "prev_avg_review_in_cnt",
    "prev_label_rate",
    "growth_rate",
    "daily_delta",
    "hit_rule_ids",
    "hit_conditions",
]
SUMMARY_COLUMNS = [
    "mach_root_label_name",
    "POC",
    "low_efficiency_strategy_count",
    "avg_review_in_cnt",
    "avg_review_done_cnt",
    "avg_label_cnt",
    "label_rate",
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_SOURCE))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--top-n", type=int, default=10)
    parser.add_argument("--sheet-url")
    parser.add_argument("--import-workbook", action="store_true")
    parser.add_argument("--send-user-id")
    parser.add_argument("--send-chat-id")
    parser.add_argument("--identity", choices=["bot", "user"], default="bot")
    parser.add_argument("--title", default="近7天低效打标策略全等级结果")
    args = parser.parse_args()

    if args.top_n <= 0:
        raise SystemExit("--top-n must be a positive integer.")
    if args.send_user_id and args.send_chat_id:
        raise SystemExit("--send-user-id and --send-chat-id cannot be used together.")

    source_path = Path(args.source)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    publish_dir = output_dir / "publish"
    publish_dir.mkdir(parents=True, exist_ok=True)

    sample = load_stage_1_sample(source_path)
    execution = sample["readonly_execution"]
    period = derive_period(sample)

    write_level_csvs(output_dir, execution)
    summary_rows = build_label_poc_summary_rows(execution["comprehensive_results"])
    write_summary_csv(output_dir / "汇总统计.csv", summary_rows)
    workbook_path = write_workbook(output_dir, execution, period, summary_rows)

    sheet_url = args.sheet_url
    if args.import_workbook and not sheet_url:
        sheet_url = import_workbook(
            workbook_path,
            name=f"低效打标全等级结果-{period['current_start']}-{period['current_end']}",
        )

    summary = build_summary(
        source_path=source_path,
        output_dir=output_dir,
        workbook_path=workbook_path,
        sample=sample,
        period=period,
        sheet_url=sheet_url,
        top_n=args.top_n,
        self_send_requested=args.send_user_id is not None or args.send_chat_id is not None,
    )
    summary["label_poc_summary_count"] = len(summary_rows)
    level_top_rows = build_level_top_rows(execution["level_results"], args.top_n)
    hash_rows = flatten_level_top_rows(level_top_rows)
    card_with_meta = render_grading_card(
        summary=summary,
        level_top_rows=level_top_rows,
        sheet_url=sheet_url,
        title=args.title,
    )
    verify_card_hash(card_with_meta, hash_rows)
    card_json = strip_internal_keys(card_with_meta)

    card_path = publish_dir / f"{REPORT_TYPE}.card.json"
    card_with_meta_path = publish_dir / f"{REPORT_TYPE}.card.with_meta.json"
    hash_check_path = publish_dir / "card_hash_check.json"
    summary_path = output_dir / "summary.json"
    publish_summary_path = publish_dir / f"{REPORT_TYPE}.publish_summary.json"
    poc_routing_path = output_dir / "poc_routing_plan.json"
    notification_draft_path = output_dir / "notification_draft.json"
    send_plan_path = output_dir / "send_plan.json"

    write_json(card_path, card_json, compact=True)
    write_json(card_with_meta_path, card_with_meta)
    write_json(
        hash_check_path,
        {
            "ok": True,
            "data_hash": card_with_meta["_meta"]["_data_hash"],
            "top_rows_count": len(hash_rows),
            "level_top_rows_count": {
                level: len(rows) for level, rows in level_top_rows.items()
            },
            "internal_meta_removed": "_meta" not in card_json,
            "design_check": card_design_check(card_with_meta),
        },
    )

    sent_payload = None
    if args.send_user_id:
        sent_payload = send_card(
            card_path=card_path,
            user_id=args.send_user_id,
            chat_id=None,
            identity=args.identity,
            idempotency_key=safe_idempotency_key(f"{REPORT_TYPE}-{OUTPUT_DATE}-{args.send_user_id}"),
        )
    elif args.send_chat_id:
        sent_payload = send_card(
            card_path=card_path,
            user_id=None,
            chat_id=args.send_chat_id,
            identity=args.identity,
            idempotency_key=safe_idempotency_key(f"{REPORT_TYPE}-{OUTPUT_DATE}-{args.send_chat_id}"),
        )

    publish_summary = {
        "report_type": REPORT_TYPE,
        "scenario_key": SCENARIO_KEY,
        "source_stage_1_result": relative_to_root(source_path),
        "output_dir": relative_to_root(output_dir),
        "summary_json": relative_to_root(summary_path),
        "workbook": relative_to_root(workbook_path),
        "summary_by_label_poc_csv": relative_to_root(output_dir / "汇总统计.csv"),
        "sheet_url": sheet_url,
        "card_json": relative_to_root(card_path),
        "card_json_with_meta": relative_to_root(card_with_meta_path),
        "card_hash_check": relative_to_root(hash_check_path),
        "sent": sent_payload is not None,
        "send_identity": args.identity if sent_payload is not None else None,
        "target_type": "user" if args.send_user_id else ("chat" if args.send_chat_id else None),
        "target_user": "self" if args.send_user_id and sent_payload is not None else None,
        "target_user_open_id_prefix": mask_identifier(args.send_user_id)
        if args.send_user_id and sent_payload is not None
        else None,
        "target_chat_id": args.send_chat_id if sent_payload is not None else None,
        "message_id": extract_message_id(sent_payload),
        "send_result": sanitize_send_payload(sent_payload),
        "notification_draft": relative_to_root(notification_draft_path),
        "send_plan": relative_to_root(send_plan_path),
    }
    summary["outputs"]["poc_routing_plan"] = "poc_routing_plan.json"
    summary["outputs"]["notification_draft"] = "notification_draft.json"
    summary["outputs"]["send_plan"] = "send_plan.json"
    summary["outputs"]["summary_by_label_poc_csv"] = "汇总统计.csv"
    summary["publish"] = publish_summary
    notification_draft = build_notification_draft(
        summary=summary,
        poc_routing_path=poc_routing_path,
        card_path=card_path,
        card_with_meta_path=card_with_meta_path,
        hash_check_path=hash_check_path,
        publish_summary=publish_summary,
    )
    send_plan = build_send_plan(
        identity=args.identity,
        publish_summary=publish_summary,
        poc_routing_path=poc_routing_path,
        card_path=card_path,
        notification_draft_path=notification_draft_path,
    )
    write_json(notification_draft_path, notification_draft)
    write_json(send_plan_path, send_plan)
    write_json(summary_path, summary)
    write_json(publish_summary_path, publish_summary)

    print(
        "Stage 2 label-rate notification draft wrote "
        f"{relative_to_root(output_dir)}; sent={publish_summary['sent']}; "
        f"message_id={publish_summary['message_id']}"
    )


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def load_stage_1_sample(path: Path) -> dict[str, Any]:
    records = load_jsonl(path)
    sample = next(
        (record for record in records if record.get("record_type") == "sample"),
        None,
    )
    if not sample:
        raise ValueError("Missing sample record in stage 1 results.")
    if sample.get("analysis_mode") != "low_label_rate_grading":
        raise ValueError("Stage 1 source must be low_label_rate_grading.")
    return sample


def derive_period(sample: dict[str, Any]) -> dict[str, str]:
    freshness = sample.get("source_footer", {}).get("data_freshness", "")
    match = re.search(r"checked_at=([^;\\s]+)", freshness)
    if match:
        checked_at = datetime.fromisoformat(match.group(1))
    else:
        checked_at = datetime.now()
    current_end = checked_at.date() - timedelta(days=1)
    current_start = current_end - timedelta(days=6)
    history_start = current_end - timedelta(days=27)
    return {
        "current_start": current_start.isoformat(),
        "current_end": current_end.isoformat(),
        "history_start": history_start.isoformat(),
        "checked_at": checked_at.isoformat(),
    }


def write_level_csvs(output_dir: Path, execution: dict[str, Any]) -> None:
    for level in LEVELS:
        rows = execution["level_results"][level]["rows"]
        write_csv(output_dir / f"{level}.csv", rows)
    write_csv(output_dir / "综合.csv", execution["comprehensive_results"])


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: csv_value(csv_column_value(row, column)) for column in CSV_COLUMNS})


def write_summary_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=SUMMARY_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {column: csv_value(row.get(column)) for column in SUMMARY_COLUMNS}
            )


def csv_value(value: Any) -> Any:
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    return value


def csv_column_value(row: dict[str, Any], column: str) -> Any:
    if column == "POC":
        return row.get("POC") or row.get("poc_name")
    return row.get(column)


def write_workbook(
    output_dir: Path,
    execution: dict[str, Any],
    period: dict[str, str],
    summary_rows: list[dict[str, Any]],
) -> Path:
    workbook_path = (
        output_dir
        / f"low_label_rate_grading_{period['current_start']}_{period['current_end']}.xlsx"
    )
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, rows in [
        ("P0", execution["level_results"]["P0"]["rows"]),
        ("P1", execution["level_results"]["P1"]["rows"]),
        ("P2", execution["level_results"]["P2"]["rows"]),
        ("Notice", execution["level_results"]["notice"]["rows"]),
        ("综合", execution["comprehensive_results"]),
    ]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(CSV_COLUMNS)
        for row in rows:
            sheet.append([csv_value(csv_column_value(row, column)) for column in CSV_COLUMNS])
        style_sheet_header(sheet)
    summary_sheet = workbook.create_sheet("汇总统计")
    summary_sheet.append(SUMMARY_COLUMNS)
    for row in summary_rows:
        summary_sheet.append([csv_value(row.get(column)) for column in SUMMARY_COLUMNS])
    style_sheet_header(summary_sheet)
    workbook.save(workbook_path)
    return workbook_path


def build_label_poc_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            str(row.get("mach_root_label_name", "")),
            str(row.get("POC") or row.get("poc_name") or "未映射"),
        )
        bucket = grouped.setdefault(
            key,
            {
                "mach_root_label_name": key[0],
                "POC": key[1],
                "_strategy_names": set(),
                "_total_review_in_cnt": 0.0,
                "_total_review_done_cnt": 0.0,
                "_total_label_cnt": 0.0,
                "_avg_review_in_cnt": 0.0,
                "_avg_review_done_cnt": 0.0,
                "_avg_label_cnt": 0.0,
            },
        )
        strategy_name = str(row.get("strategy_name") or "")
        if strategy_name:
            bucket["_strategy_names"].add(strategy_name)
        bucket["_total_review_in_cnt"] += float(row.get("total_review_in_cnt", 0) or 0)
        bucket["_total_review_done_cnt"] += float(
            row.get("total_review_done_cnt", 0) or 0
        )
        bucket["_total_label_cnt"] += float(row.get("total_label_cnt", 0) or 0)
        bucket["_avg_review_in_cnt"] += float(row.get("avg_review_in_cnt", 0) or 0)
        bucket["_avg_review_done_cnt"] += float(row.get("avg_review_done_cnt", 0) or 0)
        bucket["_avg_label_cnt"] += float(row.get("avg_label_cnt", 0) or 0)

    result: list[dict[str, Any]] = []
    for bucket in grouped.values():
        total_done = bucket["_total_review_done_cnt"]
        result.append(
            {
                "mach_root_label_name": bucket["mach_root_label_name"],
                "POC": bucket["POC"],
                "low_efficiency_strategy_count": len(bucket["_strategy_names"]),
                "avg_review_in_cnt": round(bucket["_avg_review_in_cnt"]),
                "avg_review_done_cnt": round(bucket["_avg_review_done_cnt"]),
                "avg_label_cnt": round(bucket["_avg_label_cnt"]),
                "label_rate": bucket["_total_label_cnt"] / total_done if total_done else 0.0,
            }
        )
    return sorted(
        result,
        key=lambda item: (
            -float(item["avg_review_done_cnt"]),
            str(item["mach_root_label_name"]),
            str(item["POC"]),
        ),
    )


def style_sheet_header(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="D9E8FF")
    font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"


def build_summary(
    *,
    source_path: Path,
    output_dir: Path,
    workbook_path: Path,
    sample: dict[str, Any],
    period: dict[str, str],
    sheet_url: str | None,
    top_n: int,
    self_send_requested: bool,
) -> dict[str, Any]:
    execution = sample["readonly_execution"]
    provenance = sample["provenance"]
    return {
        "schema_version": "stage_2_notification_draft.v1",
        "report_type": REPORT_TYPE,
        "scenario_key": SCENARIO_KEY,
        "run_mode": "debug_only_with_explicit_self_send"
        if self_send_requested
        else "draft_only_with_name_level_poc_routing",
        "source_stage_1_result": relative_to_root(source_path),
        "output_dir": relative_to_root(output_dir),
        "dataset_id": provenance["dataset_id"],
        "region": provenance["region"],
        "period": period,
        "level_counts": execution["level_counts"],
        "comprehensive_reason_count": execution["row_count"],
        "comprehensive_strategy_group_count": execution["row_count"],
        "fallback_reason": sample["QueryPlan"]["fallback_reason"],
        "metric_formula": execution["metric_formula"],
        "top_n": top_n,
        "sheet_url": sheet_url,
        "outputs": {
            "summary_json": "summary.json",
            "notice_csv": "notice.csv",
            "P2_csv": "P2.csv",
            "P1_csv": "P1.csv",
            "P0_csv": "P0.csv",
            "comprehensive_csv": "综合.csv",
            "workbook": workbook_path.name,
        },
        "source_footer": sample["source_footer"],
    }


def build_notification_draft(
    *,
    summary: dict[str, Any],
    poc_routing_path: Path,
    card_path: Path,
    card_with_meta_path: Path,
    hash_check_path: Path,
    publish_summary: dict[str, Any],
) -> dict[str, Any]:
    if not poc_routing_path.exists():
        raise FileNotFoundError(
            "Missing poc_routing_plan.json. Run run_stage_2_label_rate_poc_routing.py first."
        )
    poc_routing = load_json(poc_routing_path)
    constraints = poc_routing.get("routing_constraints", {})
    return {
        "schema_version": "stage_2_notification_draft_detail.v1",
        "scenario_key": SCENARIO_KEY,
        "report_type": REPORT_TYPE,
        "draft_mode": "self_preview_with_name_level_poc_routing",
        "default_self_validation": True,
        "real_poc_mapping_used": poc_routing.get("real_poc_mapping_used", False),
        "source_stage_1_result": summary["source_stage_1_result"],
        "level_counts": summary["level_counts"],
        "comprehensive_reason_count": summary["comprehensive_reason_count"],
        "comprehensive_strategy_group_count": summary[
            "comprehensive_strategy_group_count"
        ],
        "data_link": {
            "sheet_url": summary.get("sheet_url"),
            "workbook": summary["outputs"].get("workbook"),
            "csv_files": {
                "summary_by_label_poc": summary["outputs"].get(
                    "summary_by_label_poc_csv"
                ),
                "notice": summary["outputs"].get("notice_csv"),
                "P2": summary["outputs"].get("P2_csv"),
                "P1": summary["outputs"].get("P1_csv"),
                "P0": summary["outputs"].get("P0_csv"),
                "comprehensive": summary["outputs"].get("comprehensive_csv"),
            },
        },
        "card_draft": {
            "card_json": relative_to_root(card_path),
            "card_json_with_meta": relative_to_root(card_with_meta_path),
            "card_hash_check": relative_to_root(hash_check_path),
            "send_card_meta_removed": True,
        },
        "poc_routing": {
            "poc_routing_plan": relative_to_root(poc_routing_path),
            "routing_mode": poc_routing.get("routing_mode"),
            "fallback_to_default_user": poc_routing.get("fallback_to_default_user"),
            "default_recipient": poc_routing.get("default_recipient"),
            "routing_key": poc_routing.get("routing_key"),
            "contact_resolution_status": poc_routing.get("contact_resolution_status"),
            "mapped_row_count": poc_routing.get("mapped_row_count"),
            "unmapped_row_count": poc_routing.get("unmapped_row_count"),
            "missing_route_dimension_count": poc_routing.get(
                "missing_route_dimension_count"
            ),
            "poc_summary": poc_routing.get("poc_summary", []),
            "routing_rules": summarize_routing_rules(poc_routing),
            "routing_constraints": constraints,
        },
        "methodology": {
            "metric_formula": summary.get("metric_formula"),
            "period": summary.get("period"),
            "source_footer": summary.get("source_footer"),
        },
        "send_safety": {
            "current_stage": "draft_only_for_group_send",
            "current_preview_recipient": "self",
            "requires_confirmation_before_group_send": True,
            "group_send_blocked": constraints.get("group_send_blocked", True),
            "group_send_allowed": constraints.get("group_send_allowed", False),
            "sent": False,
            "real_group_send_executed": False,
            "online_write_executed": constraints.get("online_write_executed", False),
            "self_preview_sent": publish_summary.get("sent", False),
            "self_preview_message_id": publish_summary.get("message_id"),
        },
        "provenance": {
            "dataset_id": summary.get("dataset_id"),
            "region": summary.get("region"),
            "query_plan_id": summary.get("source_footer", {})
            .get("query_context", {})
            .get("query_plan_id"),
        },
    }


def summarize_routing_rules(poc_routing: dict[str, Any]) -> dict[str, Any]:
    rules = poc_routing.get("routing_rules", {})
    return {
        level: {
            "target_roles": rule.get("target_roles", []),
            "action_required": rule.get("action_required"),
            "default_recipient": rule.get("default_recipient"),
            "recipient_resolution": rule.get("recipient_resolution", {}),
            "requires_human_confirmation_before_real_send": rule.get(
                "requires_human_confirmation_before_real_send"
            ),
            "group_send_blocked": rule.get("group_send_blocked"),
            "reason_count": rule.get("reason_count"),
            "strategy_group_count": rule.get("strategy_group_count"),
            "poc_names": rule.get("poc_names", []),
            "unmapped_labels": rule.get("unmapped_labels", []),
        }
        for level, rule in rules.items()
    }


def build_send_plan(
    *,
    identity: str,
    publish_summary: dict[str, Any],
    poc_routing_path: Path,
    card_path: Path,
    notification_draft_path: Path,
) -> dict[str, Any]:
    return {
        "schema_version": "stage_2_send_plan.v1",
        "scenario_key": SCENARIO_KEY,
        "report_type": REPORT_TYPE,
        "plan_mode": "manual_confirmation_required",
        "target_type": "future_group_or_name_level_poc_after_confirmation",
        "target_source": relative_to_root(poc_routing_path),
        "content_source": {
            "card_json": relative_to_root(card_path),
            "notification_draft": relative_to_root(notification_draft_path),
        },
        "send_identity": identity,
        "requires_confirmation": True,
        "confirmation_status": "not_requested",
        "group_send_blocked": True,
        "group_send_allowed": False,
        "group_recipients": [],
        "sent": False,
        "real_group_send_executed": False,
        "online_write_executed": False,
        "blocked_reason": (
            "Stage 2 has name-level POC routing only. Feishu open_id resolution, "
            "target chat confirmation, and explicit confirmation are required before "
            "real POC/group sending."
        ),
        "self_preview": {
            "default_recipient": "self",
            "sent": publish_summary.get("sent", False),
            "message_id": publish_summary.get("message_id"),
        },
        "required_before_real_send": [
            "resolve POC names to confirmed Feishu open_id",
            "confirm target chat or POC recipients",
            "confirm send identity and content",
            "run validator with group-send gate enabled",
        ],
    }


def build_level_top_rows(
    level_results: dict[str, dict[str, Any]],
    top_n: int,
) -> dict[str, list[dict[str, Any]]]:
    return {
        level: build_top_rows(level_results.get(level, {}).get("rows", []), top_n)
        for level in ("P0", "P1", "P2", "notice")
    }


def flatten_level_top_rows(
    level_top_rows: dict[str, list[dict[str, Any]]]
) -> list[dict[str, Any]]:
    flattened: list[dict[str, Any]] = []
    for level in ("P0", "P1", "P2", "notice"):
        flattened.extend(level_top_rows.get(level, []))
    return flattened


def build_top_rows(rows: list[dict[str, Any]], top_n: int) -> list[dict[str, Any]]:
    top_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows[:top_n], 1):
        level = row["severity_level"]
        top_rows.append(
            {
                "rank": index,
                "level": [{"text": level, "color": LEVEL_COLORS.get(level, "blue")}],
                "poc_name": row.get("POC") or row.get("poc_name", "未映射"),
                "mach_root_label_name": row.get("mach_root_label_name", ""),
                "strategy_id": row.get("strategy_id", ""),
                "strategy_name": row.get("strategy_name", ""),
                "reason": row["reason"],
                "avg_in": round(float(row["avg_review_in_cnt"])),
                "avg_done": round(float(row["avg_review_done_cnt"])),
                "avg_labeled": round(float(row["avg_label_cnt"])),
                "label_rate": float(row["label_rate"]),
            }
        )
    return top_rows


def import_workbook(workbook_path: Path, name: str) -> str:
    payload = run_lark_cli(
        [
            "lark-cli",
            "sheets",
            "+workbook-import",
            "--json",
            "--as",
            "user",
            "--file",
            relative_to_cwd(workbook_path),
            "--name",
            name,
        ]
    )
    data = payload.get("data", {})
    url = data.get("url")
    if not url:
        raise RuntimeError(f"workbook import did not return url: {payload}")
    return str(url)


def send_card(
    *,
    card_path: Path,
    user_id: str | None,
    chat_id: str | None,
    identity: str,
    idempotency_key: str,
) -> dict[str, Any]:
    if bool(user_id) == bool(chat_id):
        raise ValueError("Exactly one of user_id or chat_id is required.")
    target_args = ["--user-id", user_id] if user_id else ["--chat-id", str(chat_id)]
    return run_lark_cli(
        [
            "lark-cli",
            "im",
            "+messages-send",
            "--json",
            "--as",
            identity,
            *target_args,
            "--msg-type",
            "interactive",
            "--content",
            card_path.read_text(encoding="utf-8"),
            "--idempotency-key",
            idempotency_key,
        ]
    )


def run_lark_cli(args: list[str]) -> dict[str, Any]:
    completed = subprocess.run(
        args,
        check=False,
        capture_output=True,
        text=True,
        cwd=ROOT.parent,
        env={
            **os.environ,
            "LARKSUITE_CLI_NO_UPDATE_NOTIFIER": "1",
            "LARKSUITE_CLI_NO_SKILLS_NOTIFIER": "1",
        },
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "lark-cli command failed:\n"
            f"args={args}\nstdout={completed.stdout}\nstderr={completed.stderr}"
        )
    try:
        return json.loads(completed.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"lark-cli returned non-json output: {completed.stdout}") from exc


def write_json(path: Path, value: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = (
        json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        if compact
        else json.dumps(value, ensure_ascii=False, indent=2)
    )
    path.write_text(text + "\n", encoding="utf-8")


def relative_to_root(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return str(resolved)


def relative_to_cwd(path: Path) -> str:
    resolved = path.resolve()
    cwd = ROOT.parent.resolve()
    try:
        return str(resolved.relative_to(cwd))
    except ValueError:
        return str(path)


def safe_idempotency_key(raw: str) -> str:
    key = re.sub(r"[^A-Za-z0-9-]+", "-", raw).strip("-")
    key = re.sub(r"-{2,}", "-", key)
    return key[-50:] or "label-rate-card"


def mask_identifier(value: str | None) -> str | None:
    if not value:
        return None
    return value[:10] + "..."


def extract_message_id(payload: dict[str, Any] | None) -> str | None:
    if not payload:
        return None
    data = payload.get("data", {})
    if isinstance(data, dict):
        return data.get("message_id")
    return None


def sanitize_send_payload(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not payload:
        return None
    data = payload.get("data", {})
    return {
        "ok": payload.get("ok"),
        "identity": payload.get("identity"),
        "message_id": data.get("message_id") if isinstance(data, dict) else None,
        "create_time": data.get("create_time") if isinstance(data, dict) else None,
    }


if __name__ == "__main__":
    main()
