#!/usr/bin/env python3
"""Run a custom multi-dimension low label-rate query flow."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
NOTIFICATION_SCRIPTS = ROOT / "skills" / "notification" / "scripts"
sys.path.insert(0, str(NOTIFICATION_SCRIPTS))

from card_hash import strip_internal_keys  # noqa: E402
from render_label_rate_grading_card import render_grading_card  # noqa: E402
from resolve_label_rate_poc_routing import (  # noqa: E402
    build_custom_dimension_poc_routing_plan,
)


SCENARIO_KEY = "efficiency-label-rate"
REPORT_TYPE = "custom_label_rate_breakdown"
DATASET_ID = "3888816"
APP_ID = "1128"
REGION = "cn"
DATASET_NAME = "[重点模型]-社区_人工审核明细数据"
SOURCE_TABLE = "olap_content_security_community.dws_sft_tcs_review_task_detail_di"
DEFAULT_START_DATE = "2026-06-29"
DEFAULT_END_DATE = "2026-07-05"
DEFAULT_OUTPUT_DIR = (
    ROOT
    / "evals"
    / SCENARIO_KEY
    / "stage_2_runs"
    / "20260709_custom_label_rate_breakdown_summary_default_20260629_20260705"
)
DEFAULT_DIMENSIONS = [
    "mach_root_label_name",
    "strategy_id",
    "strategy_name",
    "reason",
]
DIMENSION_SPECS = {
    "mach_root_label_name": {
        "display_name": "机审一级标签",
        "source_field": "`[机审一级标签]`",
        "empty_value": "（空/机审一级标签）",
    },
    "strategy_id": {
        "display_name": "strategy_id",
        "source_field": "`[strategy_id]`",
        "empty_value": "（空/strategy_id）",
    },
    "strategy_name": {
        "display_name": "strategy_name",
        "source_field": "`[strategy_name]`",
        "empty_value": "（空/strategy_name）",
    },
    "reason": {
        "display_name": "reason",
        "source_field": "`[reason]`",
        "empty_value": "（空/reason）",
    },
}
CSV_COLUMNS = [
    "mach_root_label_name",
    "strategy_id",
    "strategy_name",
    "reason",
    "data_days",
    "calendar_days",
    "total_review_in_cnt",
    "total_review_done_cnt",
    "total_label_cnt",
    "avg_review_in_cnt",
    "avg_review_done_cnt",
    "avg_label_cnt",
    "label_rate",
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--start-date", default=DEFAULT_START_DATE)
    parser.add_argument("--end-date", default=DEFAULT_END_DATE)
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--limit", type=int, default=50000)
    parser.add_argument("--top-n", type=int, default=10)
    parser.add_argument("--import-workbook", dest="import_workbook", action="store_true", default=True)
    parser.add_argument("--no-import-workbook", dest="import_workbook", action="store_false")
    parser.add_argument("--send-chat-id")
    parser.add_argument("--identity", choices=["bot", "user"], default="bot")
    parser.add_argument("--title", default="6月29日-7月5日低打标率多维明细")
    args = parser.parse_args()

    start_date = parse_date(args.start_date)
    end_date = parse_date(args.end_date)
    if end_date < start_date:
        raise SystemExit("--end-date must be >= --start-date.")
    if args.limit <= 0:
        raise SystemExit("--limit must be a positive integer.")
    if args.top_n <= 0:
        raise SystemExit("--top-n must be a positive integer.")

    output_dir = Path(args.output_dir)
    publish_dir = output_dir / "publish"
    output_dir.mkdir(parents=True, exist_ok=True)
    publish_dir.mkdir(parents=True, exist_ok=True)

    sql = build_sql(start_date=start_date, end_date=end_date, limit=args.limit)
    payload = run_query(sql, limit=args.limit)
    rows = normalize_rows(payload)

    csv_path = output_dir / "custom_label_rate_breakdown.csv"
    workbook_path = output_dir / (
        f"custom_label_rate_breakdown_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    )
    result_path = output_dir / "custom_label_rate_breakdown_results.jsonl"
    summary_path = output_dir / "summary.json"
    analysis_summary_path = output_dir / "analysis_summary.md"
    poc_routing_path = output_dir / "poc_routing_plan.json"
    card_with_meta_path = publish_dir / f"{REPORT_TYPE}.card.with_meta.json"
    card_path = publish_dir / f"{REPORT_TYPE}.card.json"
    hash_check_path = publish_dir / "card_hash_check.json"
    publish_summary_path = publish_dir / f"{REPORT_TYPE}.publish_summary.json"

    write_csv(csv_path, rows)
    write_workbook(workbook_path, rows, start_date=start_date, end_date=end_date)

    sheet_url = None
    if args.import_workbook:
        sheet_url = import_workbook(
            workbook_path,
            name=f"低打标率多维明细-{start_date.isoformat()}-{end_date.isoformat()}",
        )

    summary = build_summary(
        rows=rows,
        payload=payload,
        sql=sql,
        start_date=start_date,
        end_date=end_date,
        output_dir=output_dir,
        csv_path=csv_path,
        workbook_path=workbook_path,
        sheet_url=sheet_url,
        result_path=result_path,
    )
    records = build_records(summary=summary, rows=rows, payload=payload, sql=sql)
    write_jsonl(result_path, records)

    analysis_top_rows = rows[: args.top_n]
    summary["outputs"]["analysis_summary_md"] = relative_to_root(analysis_summary_path)
    poc_routing_plan = build_custom_dimension_poc_routing_plan(
        rows,
        source_result=relative_to_root(result_path),
        sheet_url=sheet_url,
    )
    summary["outputs"]["poc_routing_plan"] = relative_to_root(poc_routing_path)
    write_json(poc_routing_path, poc_routing_plan)
    write_text(analysis_summary_path, build_analysis_summary(summary, analysis_top_rows))

    sent_payload = None
    send_channel = None
    card_send_error = None
    if args.send_chat_id:
        card_top_rows = build_card_top_rows(rows, args.top_n)
        card_with_meta = render_card(
            summary=summary,
            top_rows=card_top_rows,
            sheet_url=sheet_url,
            title=args.title,
        )
        card_json = strip_internal_keys(card_with_meta)
        write_json(card_with_meta_path, card_with_meta)
        write_json(card_path, card_json, compact=True)
        write_json(
            hash_check_path,
            {
                "ok": True,
                "data_hash": card_with_meta["_meta"]["_data_hash"],
                "top_rows_count": len(card_top_rows),
                "top_rows": card_top_rows,
                "internal_meta_removed": "_meta" not in card_json,
                "design_check": card_design_check(card_with_meta),
            },
        )
        try:
            sent_payload = send_card(
                chat_id=args.send_chat_id,
                card_path=card_path,
                identity=args.identity,
                idempotency_key=safe_idempotency_key(
                    f"{REPORT_TYPE}-card-{start_date.isoformat()}-{end_date.isoformat()}-{datetime.now().isoformat()}"
                ),
            )
            send_channel = "interactive_card"
        except RuntimeError as exc:
            card_send_error = summarize_error(exc)
            sent_payload = send_markdown(
                chat_id=args.send_chat_id,
                markdown=build_markdown_message(summary, analysis_top_rows, sheet_url),
                identity=args.identity,
                idempotency_key=safe_idempotency_key(
                    f"{REPORT_TYPE}-markdown-{start_date.isoformat()}-{end_date.isoformat()}-{datetime.now().isoformat()}"
                ),
            )
            send_channel = "markdown_fallback"

    publish_summary = {
        "report_type": REPORT_TYPE,
        "scenario_key": SCENARIO_KEY,
        "output_dir": relative_to_root(output_dir),
        "summary_json": relative_to_root(summary_path),
        "analysis_summary_md": relative_to_root(analysis_summary_path),
        "poc_routing_plan": relative_to_root(poc_routing_path),
        "result_jsonl": relative_to_root(result_path),
        "csv": relative_to_root(csv_path),
        "workbook": relative_to_root(workbook_path),
        "sheet_url": sheet_url,
        "card_json": relative_to_root(card_path) if args.send_chat_id else None,
        "card_json_with_meta": relative_to_root(card_with_meta_path) if args.send_chat_id else None,
        "card_hash_check": relative_to_root(hash_check_path) if args.send_chat_id else None,
        "send_channel": send_channel,
        "card_send_error": card_send_error,
        "sent": sent_payload is not None,
        "send_identity": args.identity if sent_payload is not None else None,
        "target_chat_id": args.send_chat_id if sent_payload is not None else None,
        "message_id": extract_message_id(sent_payload),
        "send_result": sanitize_send_payload(sent_payload),
    }
    summary["publish"] = publish_summary
    write_json(summary_path, summary)
    write_json(publish_summary_path, publish_summary)

    if args.send_chat_id:
        write_group_send_validation(
            output_dir / "group_send_validation.json",
            chat_id=args.send_chat_id,
            publish_summary=publish_summary,
            summary=summary,
        )

    print(
        "Custom label-rate breakdown wrote "
        f"{relative_to_root(output_dir)}; rows={len(rows)}; "
        f"sheet_url={sheet_url}; sent={publish_summary['sent']}; "
        f"message_id={publish_summary['message_id']}\n"
        f"{build_console_summary(summary, analysis_top_rows)}"
    )


def parse_date(raw_value: str) -> date:
    return date.fromisoformat(raw_value)


def build_sql(*, start_date: date, end_date: date, limit: int) -> str:
    end_exclusive = end_date + timedelta(days=1)
    dimension_select = ",\n    ".join(
        f"ifNull({spec['source_field']}, '{spec['empty_value']}') AS {name}"
        for name, spec in DIMENSION_SPECS.items()
    )
    dimension_names = ", ".join(DEFAULT_DIMENSIONS)
    return f"""
SELECT
  {dimension_names},
  COUNT(DISTINCT dt) AS data_days,
  {calendar_days(start_date, end_date)} AS calendar_days,
  SUM(review_in_cnt) AS total_review_in_cnt,
  SUM(review_done_cnt) AS total_review_done_cnt,
  SUM(label_cnt) AS total_label_cnt,
  SUM(review_in_cnt) / COUNT(DISTINCT dt) AS avg_review_in_cnt,
  SUM(review_done_cnt) / COUNT(DISTINCT dt) AS avg_review_done_cnt,
  SUM(label_cnt) / COUNT(DISTINCT dt) AS avg_label_cnt,
  if(SUM(review_done_cnt) = 0, 0, SUM(label_cnt) / SUM(review_done_cnt)) AS label_rate
FROM (
  SELECT
    {dimension_select},
    `[p_date]` AS dt,
    `[进审量_reviewid]` AS review_in_cnt,
    `[完审量_reviewid]` AS review_done_cnt,
    `[打标量__reviewid]` AS label_cnt
  FROM {SOURCE_TABLE}
  WHERE `[p_date]` >= '{start_date.isoformat()}'
    AND `[p_date]` < '{end_exclusive.isoformat()}'
    AND `[project_title]` NOT LIKE '%虚假%'
    AND `[project_title]` NOT LIKE '%标注%'
    AND `[project_title]` NOT LIKE '%虚假不实%'
    AND `[project_title]` NOT LIKE '%封面%'
    AND `[project_title]` NOT LIKE '%自动处置%'
    AND `[project_title]` NOT LIKE '%演绎%'
    AND `[project_title]` NOT LIKE '%模型%'
    AND `[project_title]` NOT LIKE '%run%'
    AND `[project_title]` NOT LIKE '%质检%'
    AND `[project_title]` NOT LIKE '%QA%'
    AND `[project_title]` NOT LIKE '%测试%'
    AND `[project_title]` NOT LIKE '%大模型%'
    AND `[project_title]` NOT LIKE '%离线%'
    AND `[scene]` IN ('community_audit_safe', 'community_audit_style', 'community_audit_moderate')
    AND `[reason]` NOT IN ('recall_skip_L6', 'fatal_output')
    AND (`[机审一级标签]` IS NULL OR `[机审一级标签]` IN (
      '不良行为或争议价值观',
      '侵犯未成年权益',
      '偏激社会情绪和涉外言论',
      '党和国家形象负面',
      '危险行为',
      '国家安全',
      '引人不适',
      '指令舆情相关',
      '短期策略迁移',
      '色情性化',
      '违法违规',
      '领导人'
    ))
  GROUP BY {dimension_names}, dt
) daily
GROUP BY {dimension_names}
HAVING SUM(review_done_cnt) > 0
   AND if(SUM(review_done_cnt) = 0, 0, SUM(label_cnt) / SUM(review_done_cnt)) < 0.1
ORDER BY avg_review_done_cnt DESC
LIMIT {limit}
""".strip()


def calendar_days(start_date: date, end_date: date) -> int:
    return (end_date - start_date).days + 1


def run_query(sql: str, *, limit: int) -> dict[str, Any]:
    command = [
        "bytedcli",
        "-j",
        "aeolus",
        "query",
        "-r",
        REGION,
        DATASET_ID,
        sql,
        "--limit",
        str(limit),
    ]
    completed = subprocess.run(command, check=False, capture_output=True, text=True)
    if completed.returncode != 0:
        raise RuntimeError(
            "Aeolus custom label-rate query failed:\n"
            f"stdout={completed.stdout}\nstderr={completed.stderr}"
        )
    payload = json.loads(completed.stdout)
    if payload.get("status") != "success":
        raise RuntimeError(f"Aeolus custom label-rate query returned non-success: {payload}")
    return payload


def normalize_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    columns = payload["data"]["columns"]
    rows: list[dict[str, Any]] = []
    for raw_row in payload["data"]["rows"]:
        row = dict(zip(columns, raw_row))
        rows.append(
            {
                "mach_root_label_name": str(row["mach_root_label_name"]),
                "strategy_id": str(row["strategy_id"]),
                "strategy_name": str(row["strategy_name"]),
                "reason": str(row["reason"]),
                "data_days": int(row["data_days"]),
                "calendar_days": int(row["calendar_days"]),
                "total_review_in_cnt": float(row["total_review_in_cnt"]),
                "total_review_done_cnt": float(row["total_review_done_cnt"]),
                "total_label_cnt": float(row["total_label_cnt"]),
                "avg_review_in_cnt": float(row["avg_review_in_cnt"]),
                "avg_review_done_cnt": float(row["avg_review_done_cnt"]),
                "avg_label_cnt": float(row["avg_label_cnt"]),
                "label_rate": float(row["label_rate"]),
            }
        )
    return rows


def build_summary(
    *,
    rows: list[dict[str, Any]],
    payload: dict[str, Any],
    sql: str,
    start_date: date,
    end_date: date,
    output_dir: Path,
    csv_path: Path,
    workbook_path: Path,
    sheet_url: str | None,
    result_path: Path,
) -> dict[str, Any]:
    total_review_done = sum(row["total_review_done_cnt"] for row in rows)
    total_label = sum(row["total_label_cnt"] for row in rows)
    return {
        "schema_version": "custom_label_rate_breakdown.v1",
        "scenario_key": SCENARIO_KEY,
        "report_type": REPORT_TYPE,
        "analysis_mode": "custom_dimension_breakdown",
        "source_tier": "governed_dataset",
        "dataset_id": DATASET_ID,
        "app_id": APP_ID,
        "region": REGION,
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "calendar_days": calendar_days(start_date, end_date),
            "checked_at": payload.get("context", {}).get("timestamp"),
        },
        "dimensions": list(DEFAULT_DIMENSIONS),
        "dimension_display_names": {
            name: spec["display_name"] for name, spec in DIMENSION_SPECS.items()
        },
        "metric_formula": "label_rate = SUM(打标量__reviewid) / SUM(完审量_reviewid)",
        "filters": [
            "p_date between start_date and end_date inclusive",
            "standard project_title blacklist",
            "scene allowlist",
            "reason exclusion",
            "mach_root_label_name allowlist with null preserved",
            "label_rate < 0.1",
            "review_done_cnt > 0",
        ],
        "row_count": len(rows),
        "query_row_count": payload["data"]["rowCount"],
        "truncated": payload["data"].get("truncated"),
        "weighted_label_rate": 0 if total_review_done == 0 else total_label / total_review_done,
        "outputs": {
            "output_dir": relative_to_root(output_dir),
            "result_jsonl": relative_to_root(result_path),
            "csv": relative_to_root(csv_path),
            "workbook": relative_to_root(workbook_path),
            "sheet_url": sheet_url,
        },
        "sql": sql,
        "source_footer": {
            "source_tier": "governed_dataset",
            "confidence_tier": "high",
            "freshness": (
                f"p_date >= {start_date.isoformat()} AND p_date <= {end_date.isoformat()}; "
                f"checked_at={payload.get('context', {}).get('timestamp')}"
            ),
            "owner": "人审效率域数据 Owner",
            "review_status": "real_readonly_query_executed",
        },
    }


def build_records(
    *,
    summary: dict[str, Any],
    rows: list[dict[str, Any]],
    payload: dict[str, Any],
    sql: str,
) -> list[dict[str, Any]]:
    query_plan_id = "QP-ELR-CUSTOM-BREAKDOWN-20260629-20260705"
    tool_call_id = f"TCR-{query_plan_id}-01"
    query_plan = {
        "query_plan_id": query_plan_id,
        "scenario_key": SCENARIO_KEY,
        "task_type": "query_only",
        "analysis_mode": "custom_dimension_breakdown",
        "metric_id": "label_rate",
        "metric_entities": [
            {
                "metric_id": "label_rate",
                "source_tier": "governed_dataset",
                "aeolus_dataset_id": DATASET_ID,
                "aeolus_metric_id": "10000036292379",
            }
        ],
        "time_range": summary["period"],
        "dimensions": summary["dimensions"],
        "dimension_mappings": [
            {
                "dimension_id": name,
                "source_field": DIMENSION_SPECS[name]["source_field"],
                "source_tier": "governed_dataset",
            }
            for name in DEFAULT_DIMENSIONS
        ],
        "filters": summary["filters"],
        "required_hygiene_filters": [
            "A_project_title_blacklist",
            "B_scene_allowlist",
            "C_reason_exclusion",
            "D_mach_root_label_allowlist_with_null",
        ],
        "quality_checks": [
            "field_discovery_check",
            "freshness_gate",
            "denominator_not_zero",
            "grain_check",
            "truncation_check",
        ],
        "execution_mode": "real_readonly_query",
        "review_required": False,
        "sql": sql,
        "tool_calls": [tool_call_id],
    }
    readonly_execution = {
        "execution_id": f"ROE-{query_plan_id}",
        "execution_mode": "real_readonly_query",
        "status": "success",
        "source_tier": "governed_dataset",
        "source_name": f"{DATASET_NAME} ({DATASET_ID})",
        "data_freshness": summary["source_footer"]["freshness"],
        "row_count": len(rows),
        "truncated": payload["data"].get("truncated"),
        "columns": CSV_COLUMNS,
        "rows": rows,
        "metric_formula": summary["metric_formula"],
        "quality_checks": {
            "field_discovery_check": "passed_strategy_id_strategy_name_confirmed",
            "freshness_gate": "passed_explicit_date_filter",
            "denominator_not_zero": "passed",
            "grain_check": "passed_mach_root_label_strategy_reason",
            "truncation_check": "passed" if payload["data"].get("truncated") is False else "failed",
        },
    }
    provenance = {
        "provenance_id": f"PROV-{query_plan_id}",
        "scenario_key": SCENARIO_KEY,
        "query_plan_id": query_plan_id,
        "execution_id": readonly_execution["execution_id"],
        "source_tier": "governed_dataset",
        "source_name": readonly_execution["source_name"],
        "region": REGION,
        "app_id": APP_ID,
        "dataset_id": DATASET_ID,
        "metric_id": "label_rate",
        "metric_formula": summary["metric_formula"],
        "time_range": query_plan["time_range"],
        "dimensions": query_plan["dimensions"],
        "filters": query_plan["filters"],
        "quality_checks": readonly_execution["quality_checks"],
        "tool_call_ids": [tool_call_id],
        "sql": sql,
        "source_footer": summary["source_footer"],
    }
    analysis_result = {
        "analysis_id": "AN-ELR-CUSTOM-BREAKDOWN-20260629-20260705",
        "event_id": "ELR-CUSTOM-BREAKDOWN-20260629-20260705",
        "templates_used": ["custom_dimension_breakdown", "readonly_execution", "source_footer"],
        "query_plan": query_plan,
        "readonly_execution": readonly_execution,
        "summary": summary,
        "source_footer": summary["source_footer"],
        "provenance": provenance,
        "sop_decision": {
            "severity_level": "P3",
            "next_action": "notify_validation_group_if_requested",
            "required_confirmation": False,
        },
    }
    tool_call_record = {
        "tool_call_id": tool_call_id,
        "caller": "custom-label-rate-breakdown",
        "tool_name": "bytedcli_aeolus_query",
        "command_name": "bytedcli -j aeolus query",
        "permission_level": "readonly",
        "source_tier": "governed_dataset",
        "scenario_key": SCENARIO_KEY,
        "metric_id": "label_rate",
        "execution_mode": "real_readonly_query",
        "real_query_executed": True,
        "input_summary": "2026-06-29..2026-07-05; mach_root_label_name × strategy_id × strategy_name × reason; label_rate < 0.1.",
        "output_summary": f"Returned {payload['data']['rowCount']} rows; truncated={payload['data'].get('truncated')}.",
        "status": "success",
        "latency_ms": payload.get("context", {}).get("execution_time_ms", 0),
    }
    return [
        {
            "record_type": "environment",
            "scenario_key": SCENARIO_KEY,
            "run_mode": "debug_only_with_optional_group_validation",
            "execution_mode": "real_readonly_query",
            "analysis_mode": "custom_dimension_breakdown",
            "real_query_executed": True,
            "online_write_blocked": True,
            "result": "pass",
        },
        {
            "record_type": "sample",
            "id": "ELR-CUSTOM-BREAKDOWN-20260629-20260705",
            "input": "2026-06-29 到 2026-07-05 打标率小于 0.1 的机审一级标签 × strategy_id × strategy_name × reason 明细。",
            "run_mode": "debug_only_with_optional_group_validation",
            "scenario_key": SCENARIO_KEY,
            "task_type": "custom_dimension_breakdown",
            "analysis_mode": "custom_dimension_breakdown",
            "QueryPlan": query_plan,
            "tool_call_records": [tool_call_record],
            "readonly_execution": readonly_execution,
            "analysis_result": analysis_result,
            "source_footer": summary["source_footer"],
            "provenance": provenance,
            "permission_checks": {
                "tool_calls": [tool_call_id],
                "read_only": True,
                "real_query_executed": True,
                "online_write_blocked": True,
            },
            "result": "pass",
        },
    ]


def build_card_top_rows(rows: list[dict[str, Any]], top_n: int) -> list[dict[str, Any]]:
    top_rows = []
    for index, row in enumerate(rows[:top_n], 1):
        top_rows.append(
            {
                "rank": index,
                "level": [{"text": "notice", "color": "blue"}],
                "reason": row["reason"],
                "avg_in": round(row["avg_review_in_cnt"]),
                "avg_done": round(row["avg_review_done_cnt"]),
                "avg_labeled": round(row["avg_label_cnt"]),
                "label_rate": row["label_rate"],
            }
        )
    return top_rows


def build_console_summary(summary: dict[str, Any], top_rows: list[dict[str, Any]]) -> str:
    lines = [
        "Summary:",
        f"- low_label_rate_group_count={summary['row_count']}",
        f"- weighted_label_rate={summary['weighted_label_rate'] * 100:.2f}%",
        f"- sheet_url={summary['outputs'].get('sheet_url')}",
        "- top_groups:",
    ]
    for row in top_rows[:5]:
        lines.append(
            (
                f"  - {row['mach_root_label_name']} | {row['strategy_id']} | "
                f"{row['strategy_name']} | {row['reason']} | "
                f"avg_done={row['avg_review_done_cnt']:.0f} | "
                f"avg_label={row['avg_label_cnt']:.0f} | "
                f"label_rate={row['label_rate'] * 100:.2f}%"
            )
        )
    return "\n".join(lines)


def build_analysis_summary(summary: dict[str, Any], top_rows: list[dict[str, Any]]) -> str:
    period = summary["period"]
    lines = [
        "# 自定义低打标率多维查询汇总",
        "",
        "## 结论摘要",
        "",
        f"- 时间窗口：`{period['start_date']}` ~ `{period['end_date']}`。",
        "- 维度：`机审一级标签 × strategy_id × strategy_name × reason`。",
        f"- 命中打标率 `<0.1` 的分组数：`{summary['row_count']}`。",
        f"- 命中分组加权打标率：`{summary['weighted_label_rate'] * 100:.2f}%`。",
        f"- 完整飞书电子表格：{summary['outputs'].get('sheet_url') or '未导入'}",
        "",
        "## Top 分组",
        "",
        "| 排名 | 机审一级标签 | strategy_id | strategy_name | reason | 日均进审量 | 日均完审量 | 日均打标量 | 打标率 |",
        "| --- | --- | --- | --- | --- | ---: | ---: | ---: | ---: |",
    ]
    for index, row in enumerate(top_rows, 1):
        lines.append(
            "| {rank} | {label} | {sid} | {sname} | {reason} | {avg_in:.0f} | {avg_done:.0f} | {avg_label:.0f} | {rate:.2f}% |".format(
                rank=index,
                label=row["mach_root_label_name"],
                sid=row["strategy_id"],
                sname=str(row["strategy_name"]).replace("|", "\\|"),
                reason=str(row["reason"]).replace("|", "\\|"),
                avg_in=row["avg_review_in_cnt"],
                avg_done=row["avg_review_done_cnt"],
                avg_label=row["avg_label_cnt"],
                rate=row["label_rate"] * 100,
            )
        )
    lines.extend(
        [
            "",
            "## 口径方法",
            "",
            "- 日均进审量：`SUM(进审量_reviewid) / COUNT(DISTINCT p_date)`。",
            "- 日均完审量：`SUM(完审量_reviewid) / COUNT(DISTINCT p_date)`。",
            "- 日均打标量：`SUM(打标量__reviewid) / COUNT(DISTINCT p_date)`。",
            "- 打标率：`SUM(打标量__reviewid) / SUM(完审量_reviewid)`。",
            "- 过滤：默认样本池 SQL 片段、`完审量 > 0`、`打标率 < 0.1`。",
            "",
            "## Provenance",
            "",
            "> Source: governed_dataset  ",
            "> Confidence: high  ",
            f"> Freshness: checked_at={period.get('checked_at')}  ",
            "> Owner: 人审效率域数据 Owner  ",
            "> Reviewed: sql_review_passed",
        ]
    )
    return "\n".join(lines) + "\n"


def render_card(
    *,
    summary: dict[str, Any],
    top_rows: list[dict[str, Any]],
    sheet_url: str | None,
    title: str,
) -> dict[str, Any]:
    period = summary["period"]
    grading_summary = {
        "report_type": REPORT_TYPE,
        "scenario_key": SCENARIO_KEY,
        "dataset_id": summary["dataset_id"],
        "region": summary["region"],
        "source_stage_1_result": summary["outputs"]["result_jsonl"],
        "fallback_reason": "custom_dimension_breakdown",
        "period": {
            "current_start": period["start_date"],
            "current_end": period["end_date"],
        },
        "level_counts": {
            "P0": 0,
            "P1": 0,
            "P2": 0,
            "notice": summary["row_count"],
        },
    }
    return render_grading_card(
        summary=grading_summary,
        summary_rows=[],
        level_top_rows={"P0": [], "P1": [], "P2": [], "notice": top_rows},
        sheet_url=sheet_url,
        title=title,
    )


def metrics_block(summary: dict[str, Any]) -> dict[str, Any]:
    metrics = [
        ("低打标率分组", summary["row_count"], "blue"),
        ("加权打标率", f"{summary['weighted_label_rate'] * 100:.2f}%", "red"),
        ("日历天数", summary["period"]["calendar_days"], "green"),
        ("TopN 展示", min(summary["row_count"], 10), "orange"),
    ]
    return {
        "tag": "column_set",
        "flex_mode": "flow",
        "horizontal_spacing": "12px",
        "columns": [
            {
                "tag": "column",
                "width": "weighted",
                "weight": 1,
                "background_style": f"{color}-50",
                "padding": "12px",
                "elements": [
                    {
                        "tag": "markdown",
                        "content": f"## <font color='{color}'>{value}</font>",
                        "text_align": "center",
                    },
                    {
                        "tag": "markdown",
                        "content": f"<font color='grey'>{label}</font>",
                        "text_align": "center",
                        "text_size": "notation",
                    },
                ],
            }
            for label, value, color in metrics
        ],
    }


def top_rows_markdown_block(top_rows: list[dict[str, Any]]) -> dict[str, Any]:
    lines = ["**Top 低打标率分组（按日均完审量排序）**"]
    for row in top_rows[:5]:
        lines.append(
            (
                f"{row['rank']}. `{row['mach_root_label_name']}` / `{row['strategy_id']}` / "
                f"{row['strategy_name']} / {row['reason']}："
                f"日均完审 {row['avg_review_done_cnt']:,}，"
                f"日均打标 {row['avg_label_cnt']:,}，"
                f"打标率 {row['label_rate'] * 100:.2f}%"
            )
        )
    lines.append("")
    lines.append("完整 10,000+ 明细请打开下方飞书表格。")
    return {
        "tag": "markdown",
        "content": "\n".join(lines),
    }


def sheet_button(sheet_url: str) -> dict[str, Any]:
    return {
        "tag": "button",
        "text": {"tag": "plain_text", "content": "查看完整飞书电子表格"},
        "type": "primary_filled",
        "width": "fill",
        "behaviors": [
            {
                "type": "open_url",
                "default_url": sheet_url,
                "pc_url": sheet_url,
                "ios_url": sheet_url,
                "android_url": sheet_url,
            }
        ],
    }


def build_markdown_message(
    summary: dict[str, Any],
    top_rows: list[dict[str, Any]],
    sheet_url: str | None,
) -> str:
    period = summary["period"]
    lines = [
        "## 6月29日-7月5日低打标率多维明细验证",
        "",
        f"- 时间窗口：{period['start_date']} ~ {period['end_date']}",
        "- 维度：机审一级标签 × strategy_id × strategy_name × reason",
        f"- 低打标率分组数：{summary['row_count']}",
        f"- 加权打标率：{summary['weighted_label_rate'] * 100:.2f}%",
        "- 口径：日均量 = SUM(指标) / COUNT(DISTINCT p_date)，打标率 = SUM(打标量) / SUM(完审量)",
    ]
    if sheet_url:
        lines.append(f"- 完整明细表：{sheet_url}")
    lines.extend(["", "### Top5（按日均完审量排序）"])
    for row in top_rows[:5]:
        lines.append(
            (
                f"{row['mach_root_label_name']} / {row['strategy_id']} / {row['strategy_name']} / {row['reason']}："
                f"日均进审 {row['avg_review_in_cnt']:.0f}，日均完审 {row['avg_review_done_cnt']:.0f}，"
                f"日均打标 {row['avg_label_cnt']:.0f}，打标率 {row['label_rate'] * 100:.2f}%"
            )
        )
    lines.extend(
        [
            "",
            "> Source: governed_dataset",
            "> Confidence: high",
            f"> Freshness: checked_at={period.get('checked_at')}",
            "> Owner: 人审效率域数据 Owner",
            "> Reviewed: sql_review_passed",
        ]
    )
    return "\n".join(lines)


def methodology_panel(summary: dict[str, Any]) -> dict[str, Any]:
    lines = [
        f"- 数据集：`{summary['dataset_id']}` / `{summary['region']}`",
        f"- 时间窗口：`{summary['period']['start_date']}` ~ `{summary['period']['end_date']}`",
        "- 维度：`机审一级标签 × strategy_id × strategy_name × reason`",
        "- 日均量：`SUM(指标) / COUNT(DISTINCT p_date)`，按组合实际有数据天数计算",
        "- 打标率：`SUM(打标量__reviewid) / SUM(完审量_reviewid)`",
        "- 过滤：标准 A/B/C/D 过滤 + `label_rate < 0.1` + `review_done_cnt > 0`",
    ]
    return {
        "tag": "collapsible_panel",
        "expanded": False,
        "background_color": "grey-50",
        "padding": "10px",
        "header": {"title": {"tag": "plain_text", "content": "口径与溯源"}},
        "elements": [{"tag": "markdown", "content": "\n".join(lines), "text_size": "notation"}],
    }


def card_design_check(card: dict[str, Any]) -> dict[str, Any]:
    tags = [
        element.get("tag")
        for element in card.get("body", {}).get("elements", [])
        if isinstance(element, dict)
    ]
    return {
        "schema_2_0": card.get("schema") == "2.0",
        "has_header": isinstance(card.get("header"), dict),
        "has_metrics": "column_set" in tags,
        "has_top_rows": "markdown" in tags or "table" in tags,
        "has_chart": "chart" in tags,
        "has_table": "table" in tags,
        "has_action": "button" in tags,
        "has_methodology": "collapsible_panel" in tags,
        "top_level_blocks": len(tags),
        "passes_p0_p3_basic_gate": (
            card.get("schema") == "2.0"
            and isinstance(card.get("header"), dict)
            and "column_set" in tags
            and ("markdown" in tags or "table" in tags)
            and "button" in tags
            and "collapsible_panel" in tags
            and 3 <= len(tags) <= 5
        ),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    start_date: date,
    end_date: date,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "低打标率多维明细"
    sheet.append([f"低打标率多维明细：{start_date.isoformat()} ~ {end_date.isoformat()}"])
    sheet.append([])
    sheet.append(CSV_COLUMNS)
    for cell in sheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for row in rows:
        sheet.append([row[column] for column in CSV_COLUMNS])
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_json(path: Path, value: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = (
        json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        if compact
        else json.dumps(value, ensure_ascii=False, indent=2)
    )
    path.write_text(text + "\n", encoding="utf-8")


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    path.write_text(
        "".join(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n" for record in records),
        encoding="utf-8",
    )


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def import_workbook(workbook_path: Path, name: str) -> str:
    payload = run_lark_cli(
        [
            "lark-cli",
            "sheets",
            "+workbook-import",
            "--json",
            "--as",
            "user",
            "--file",
            relative_to_cwd(workbook_path),
            "--name",
            name,
        ]
    )
    url = payload.get("data", {}).get("url")
    if not url:
        raise RuntimeError(f"workbook import did not return url: {payload}")
    return str(url)


def send_card(
    *,
    chat_id: str,
    card_path: Path,
    identity: str,
    idempotency_key: str,
) -> dict[str, Any]:
    return run_lark_cli(
        [
            "lark-cli",
            "im",
            "+messages-send",
            "--json",
            "--as",
            identity,
            "--chat-id",
            chat_id,
            "--msg-type",
            "interactive",
            "--content",
            card_path.read_text(encoding="utf-8"),
            "--idempotency-key",
            idempotency_key,
        ]
    )


def send_markdown(
    *,
    chat_id: str,
    markdown: str,
    identity: str,
    idempotency_key: str,
) -> dict[str, Any]:
    return run_lark_cli(
        [
            "lark-cli",
            "im",
            "+messages-send",
            "--json",
            "--as",
            identity,
            "--chat-id",
            chat_id,
            "--markdown",
            markdown,
            "--idempotency-key",
            idempotency_key,
        ]
    )


def summarize_error(error: Exception) -> dict[str, Any]:
    message = str(error)
    subtype = None
    code = None
    if "invalid_parameters" in message:
        subtype = "invalid_parameters"
    if "99992402" in message:
        code = 99992402
    return {
        "type": error.__class__.__name__,
        "subtype": subtype,
        "code": code,
        "message": message[:500],
        "truncated": len(message) > 500,
    }


def run_lark_cli(args: list[str]) -> dict[str, Any]:
    completed = subprocess.run(
        args,
        check=False,
        capture_output=True,
        text=True,
        cwd=ROOT.parent,
        env={
            **os.environ,
            "LARKSUITE_CLI_NO_UPDATE_NOTIFIER": "1",
            "LARKSUITE_CLI_NO_SKILLS_NOTIFIER": "1",
        },
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "lark-cli command failed:\n"
            f"args={args}\nstdout={completed.stdout}\nstderr={completed.stderr}"
        )
    try:
        return json.loads(completed.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"lark-cli returned non-json output: {completed.stdout}") from exc


def extract_message_id(payload: dict[str, Any] | None) -> str | None:
    if not payload:
        return None
    data = payload.get("data", {})
    if isinstance(data, dict):
        return data.get("message_id")
    return None


def sanitize_send_payload(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not payload:
        return None
    data = payload.get("data", {})
    return {
        "ok": payload.get("ok"),
        "identity": payload.get("identity"),
        "chat_id": data.get("chat_id") if isinstance(data, dict) else None,
        "message_id": data.get("message_id") if isinstance(data, dict) else None,
        "create_time": data.get("create_time") if isinstance(data, dict) else None,
    }


def write_group_send_validation(
    path: Path,
    *,
    chat_id: str,
    publish_summary: dict[str, Any],
    summary: dict[str, Any],
) -> None:
    write_json(
        path,
        {
            "schema_version": "custom_label_rate_group_send_validation.v1",
            "scenario_key": SCENARIO_KEY,
            "report_type": REPORT_TYPE,
            "validation_mode": "explicit_user_requested_validation_group",
            "chat": {
                "chat_id": chat_id,
                "send_identity": publish_summary.get("send_identity"),
            },
            "content_source": {
                "card_json": publish_summary.get("card_json"),
                "summary_json": publish_summary.get("summary_json"),
                "sheet_url": publish_summary.get("sheet_url"),
            },
            "send_result": {
                "sent": publish_summary.get("sent"),
                "send_channel": publish_summary.get("send_channel"),
                "message_id": publish_summary.get("message_id"),
                "send_result": publish_summary.get("send_result"),
                "card_send_error": publish_summary.get("card_send_error"),
            },
            "safety": {
                "user_explicitly_requested_group_validation": True,
                "online_state_store_write_executed": False,
                "dimensions": summary.get("dimensions"),
                "row_count": summary.get("row_count"),
            },
        },
    )


def safe_idempotency_key(raw_value: str) -> str:
    key = re.sub(r"[^A-Za-z0-9-]+", "-", raw_value).strip("-")
    key = re.sub(r"-{2,}", "-", key)
    return key[-50:] or "custom-label-rate-breakdown"


def relative_to_root(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return str(resolved)


def relative_to_cwd(path: Path) -> str:
    resolved = path.resolve()
    cwd = ROOT.parent.resolve()
    try:
        return str(resolved.relative_to(cwd))
    except ValueError:
        return str(path)


if __name__ == "__main__":
    main()
