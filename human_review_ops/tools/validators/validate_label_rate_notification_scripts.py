#!/usr/bin/env python3
"""Smoke-test reusable label-rate notification Skill scripts."""

from __future__ import annotations

import csv
import json
import sys
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
NOTIFICATION_SCRIPTS = ROOT / "skills" / "notification" / "scripts"
sys.path.insert(0, str(NOTIFICATION_SCRIPTS))

from card_hash import strip_internal_keys, verify_card_hash  # noqa: E402
import label_rate_notification_artifacts as notification_artifacts  # noqa: E402
import sheet_importer  # noqa: E402
from label_rate_notification_artifacts import (  # noqa: E402
    build_card_summary_rows,
    build_label_rate_notification_artifacts,
    build_level_top_rows,
    flatten_level_top_rows,
)
from label_rate_weekly_summary_comparison import (  # noqa: E402
    COMPARISON_SHEET_NAME,
    build_weekly_summary_comparison,
)
from render_label_rate_grading_card import card_design_check  # noqa: E402

FORBIDDEN_CARD_CONTRACT_TERMS = (
    "各等级命中 " + "reason 数柱状图",
    "Top " + "reason",
    "四维策略" + "分组",
    "送审" + "原因",
)


def make_row(
    *,
    level: str,
    label: str,
    poc: str,
    strategy_id: str,
    reason: str,
    avg_done: int,
    label_rate: float,
) -> dict[str, Any]:
    return {
        "severity_level": level,
        "severity_priority": {"P0": 0, "P1": 1, "P2": 2, "notice": 3}[level],
        "mach_root_label_name": label,
        "strategy_id": strategy_id,
        "strategy_name": f"{strategy_id}_name",
        "reason": reason,
        "POC": poc,
        "poc_name": poc,
        "poc_open_id": None,
        "poc_mapping_status": "mapped_name_only",
        "data_days": 7,
        "total_review_in_cnt": avg_done * 7,
        "total_review_done_cnt": avg_done * 7,
        "total_label_cnt": round(avg_done * 7 * label_rate),
        "avg_review_in_cnt": avg_done,
        "avg_review_done_cnt": avg_done,
        "avg_label_cnt": round(avg_done * label_rate, 2),
        "label_rate": label_rate,
        "prev_avg_review_in_cnt": avg_done - 10,
        "prev_label_rate": label_rate,
        "growth_rate": 0.25,
        "daily_delta": 10,
        "hit_rule_id": f"{level}_smoke_rule",
        "hit_rule_ids": [f"{level}_smoke_rule"],
        "hit_condition": "smoke low label-rate condition",
        "hit_conditions": ["smoke low label-rate condition"],
    }


def build_sample() -> dict[str, Any]:
    p0 = make_row(
        level="P0",
        label="国家安全",
        poc="杜衡",
        strategy_id="strategy_p0",
        reason="reason_p0",
        avg_done=12000,
        label_rate=0.01,
    )
    p1 = make_row(
        level="P1",
        label="领导人",
        poc="宋诗慧",
        strategy_id="strategy_p1",
        reason="reason_p1",
        avg_done=6000,
        label_rate=0.02,
    )
    p2 = make_row(
        level="P2",
        label="违法违规",
        poc="叶健",
        strategy_id="strategy_p2",
        reason="reason_p2",
        avg_done=3000,
        label_rate=0.04,
    )
    notice = make_row(
        level="notice",
        label="危险行为",
        poc="陈雅静",
        strategy_id="strategy_notice",
        reason="reason_notice",
        avg_done=800,
        label_rate=0.08,
    )
    notice_overlap = make_row(
        level="notice",
        label="违法违规",
        poc="叶健",
        strategy_id="strategy_notice_overlap",
        reason="reason_notice_overlap",
        avg_done=700,
        label_rate=0.07,
    )
    notice_rows = [notice, notice_overlap]
    comprehensive = [p0, p1, p2, *notice_rows]
    execution = {
        "analysis_mode": "low_label_rate_grading",
        "execution_mode": "smoke",
        "status": "success",
        "level_counts": {"notice": len(notice_rows), "P2": 1, "P1": 1, "P0": 1},
        "level_results": {
            "notice": {"rows": notice_rows},
            "P2": {"rows": [p2]},
            "P1": {"rows": [p1]},
            "P0": {"rows": [p0]},
        },
        "comprehensive_results": comprehensive,
        "row_count": len(comprehensive),
        "metric_formula": "`label_rate` = SUM(`[打标量__reviewid]`) / SUM(`[完审量_reviewid]`)",
    }
    return {
        "record_type": "sample",
        "id": "smoke-label-rate-notification",
        "scenario_key": "efficiency-label-rate",
        "task_type": "notification",
        "analysis_mode": "low_label_rate_grading",
        "run_mode": "debug_only",
        "QueryPlan": {
            "query_plan_id": "QP-SMOKE-LABEL-RATE-NOTIFICATION",
            "scenario_key": "efficiency-label-rate",
            "task_type": "low_label_rate_grading",
            "fallback_reason": "smoke_test_fixture",
        },
        "readonly_execution": execution,
        "provenance": {
            "dataset_id": "smoke_dataset",
            "region": "cn",
            "query_plan_id": "QP-SMOKE-LABEL-RATE-NOTIFICATION",
        },
        "source_footer": {
            "source_tier": "governed_dataset",
            "metric_definition_version": "smoke",
            "data_freshness": "p_date >= today() - 7 AND p_date < today(); checked_at=2026-07-09T12:00:00+08:00",
            "owner": "人审效率域数据 Owner",
            "confidence_tier": "high",
            "review_status": "smoke_fixture",
            "scenario_key": "efficiency-label-rate",
            "metric_id": "label_rate",
            "quality_checks": ["smoke_fixture"],
        },
    }


def write_source(path: Path) -> None:
    sample = build_sample()
    path.write_text(json.dumps(sample, ensure_ascii=False) + "\n", encoding="utf-8")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def validate_artifacts(artifacts: Any) -> None:
    output_dir = artifacts.output_dir
    required = [
        "summary.json",
        "notification_draft.json",
        "send_plan.json",
        "poc_routing_plan.json",
        "汇总统计.csv",
        "notice.csv",
        "P2.csv",
        "P1.csv",
        "P0.csv",
        "综合.csv",
        "综合_剔除+1同意.csv",
        "汇总统计_剔除+1同意.csv",
        "publish/low_efficiency_grading.card.json",
        "publish/low_efficiency_grading.card.with_meta.json",
        "publish/card_hash_check.json",
        "publish/low_efficiency_grading.publish_summary.json",
    ]
    for relative in required:
        if not (output_dir / relative).exists():
            raise AssertionError(f"Missing smoke artifact: {relative}")

    summary = read_json(output_dir / "summary.json")
    notification_draft = read_json(output_dir / "notification_draft.json")
    send_plan = read_json(output_dir / "send_plan.json")
    poc_routing = read_json(output_dir / "poc_routing_plan.json")
    card = read_json(output_dir / "publish" / "low_efficiency_grading.card.json")
    card_with_meta = read_json(
        output_dir / "publish" / "low_efficiency_grading.card.with_meta.json"
    )
    hash_check = read_json(output_dir / "publish" / "card_hash_check.json")

    if summary.get("schema_version") != "stage_2_notification_draft.v1":
        raise AssertionError("summary schema mismatch.")
    if summary.get("outputs", {}).get("summary_by_label_poc_csv") != "汇总统计.csv":
        raise AssertionError("summary must expose summary_by_label_poc_csv.")
    if summary.get("label_poc_summary_count") != 2:
        raise AssertionError("summary label_poc_summary_count mismatch.")
    if notification_draft.get("real_poc_mapping_used") is not True:
        raise AssertionError("notification_draft must use name-level POC mapping.")
    if notification_draft.get("send_safety", {}).get("group_send_blocked") is not True:
        raise AssertionError("notification_draft must block group send.")
    if send_plan.get("requires_confirmation") is not True:
        raise AssertionError("send_plan must require confirmation.")
    if send_plan.get("sent") is not False:
        raise AssertionError("send_plan must not mark real send as sent.")
    if poc_routing.get("routing_mode") != "mach_root_label_mapping":
        raise AssertionError("poc_routing must use mach_root_label_mapping.")
    if poc_routing.get("mapped_row_count") != 5:
        raise AssertionError("poc_routing mapped_row_count mismatch.")

    summary_rows = read_csv_rows(output_dir / "汇总统计.csv")
    if len(summary_rows) != 2:
        raise AssertionError("summary CSV row count mismatch.")
    for field in ("机审一级标签", "POC", "低效策略打标率"):
        if field not in summary_rows[0]:
            raise AssertionError(f"summary CSV missing field: {field}")
    for row in summary_rows:
        done = float(row.get("低效策略日均完审量") or 0)
        label = float(row.get("低效策略日均打标量") or 0)
        rate = float(row.get("低效策略打标率") or 0)
        expected_rate = label / done if done else 0.0
        if abs(rate - expected_rate) > 1e-12:
            raise AssertionError("summary CSV label_rate must equal label / done.")
    comprehensive_rows = read_csv_rows(output_dir / "综合.csv")
    if len(comprehensive_rows) != 5:
        raise AssertionError("comprehensive CSV row count mismatch.")
    if "预警等级" not in comprehensive_rows[0]:
        raise AssertionError("comprehensive CSV must keep severity level.")
    notice_rows = read_csv_rows(output_dir / "notice.csv")
    expected_summary_counts = expected_summary_counts_from_notice(notice_rows)
    actual_summary_counts = {
        (row.get("机审一级标签", ""), row.get("POC", "")): int(
            row.get("低效策略数") or 0
        )
        for row in summary_rows
    }
    if actual_summary_counts != expected_summary_counts:
        raise AssertionError("summary CSV must aggregate Notice rows by label and POC.")
    plus1_period_field = "+1同意日期是否在本次统计周期前"
    if plus1_period_field not in comprehensive_rows[0]:
        raise AssertionError(f"comprehensive CSV missing field: {plus1_period_field}")
    if {row.get(plus1_period_field) for row in comprehensive_rows} - {"是", "否"}:
        raise AssertionError("comprehensive CSV plus1 period flag must be 是/否.")
    if len(read_csv_rows(output_dir / "P0.csv")) != 1:
        raise AssertionError("P0 CSV row count mismatch.")

    workbook_files = list(output_dir.glob("low_label_rate_grading_*.xlsx"))
    if len(workbook_files) != 1:
        raise AssertionError("expected exactly one smoke workbook.")
    workbook = load_workbook(workbook_files[0], read_only=True)
    try:
        expected_sheets = {
            "P0",
            "P1",
            "P2",
            "Notice",
            "综合",
            "综合_剔除+1同意",
            "汇总统计",
            "汇总统计_剔除+1同意",
        }
        if set(workbook.sheetnames) != expected_sheets:
            raise AssertionError("workbook sheet structure mismatch.")
    finally:
        workbook.close()

    if "_meta" in card:
        raise AssertionError("send card must not contain _meta.")
    if strip_internal_keys(card_with_meta) != card:
        raise AssertionError("card must equal stripped card_with_meta.")
    verify_card_hash(card_with_meta, card_with_meta["_meta"]["hash_input"])
    if hash_check.get("ok") is not True:
        raise AssertionError("hash_check must be ok.")
    design_check = card_design_check(card_with_meta)
    assert_no_old_card_contract(card_with_meta, design_check)
    if design_check.get("passes_p0_p3_basic_gate") is not True:
        raise AssertionError("card design smoke gate failed.")


def expected_summary_counts_from_notice(
    notice_rows: list[dict[str, str]],
) -> dict[tuple[str, str], int]:
    grouped: dict[tuple[str, str], set[str]] = {}
    for row in notice_rows:
        key = (row.get("机审一级标签", ""), row.get("POC", "") or "未映射")
        strategy_key = (row.get("策略ID") or row.get("策略名称") or "").strip()
        if not strategy_key:
            continue
        grouped.setdefault(key, set()).add(strategy_key)
    return {key: len(values) for key, values in grouped.items()}


def iter_card_tags(value: Any) -> list[str]:
    tags: list[str] = []
    if isinstance(value, dict):
        tag = value.get("tag")
        if isinstance(tag, str):
            tags.append(tag)
        for child in value.values():
            tags.extend(iter_card_tags(child))
    elif isinstance(value, list):
        for item in value:
            tags.extend(iter_card_tags(item))
    return tags


def assert_no_old_card_contract(
    card: dict[str, Any],
    design_check: dict[str, Any],
) -> None:
    rendered = json.dumps(card, ensure_ascii=False)
    for term in FORBIDDEN_CARD_CONTRACT_TERMS:
        if term in rendered:
            raise AssertionError(f"Card still contains old contract term: {term}")
    if "chart" in iter_card_tags(card):
        raise AssertionError("Card must not contain chart elements.")
    if "has_chart" in design_check:
        raise AssertionError("card_design_check must not expose has_chart.")


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="label-rate-notification-smoke-") as tmp:
        tmp_path = Path(tmp)
        source_path = tmp_path / "source.jsonl"
        output_dir = tmp_path / "output"
        write_source(source_path)
        artifacts = build_label_rate_notification_artifacts(
            source_path=source_path,
            output_dir=output_dir,
            top_n=2,
            sheet_url="https://example.com/sheets/smoke",
            identity="bot",
            title="近7天低效打标策略全等级结果",
            self_send_requested=False,
            sent_payload=None,
            target_user_id=None,
            target_chat_id=None,
        )
        validate_artifacts(artifacts)
    validate_sheet_import_optin()
    validate_weekly_filtered_summary_comparison()
    print("Label-rate notification scripts smoke OK.")


def validate_sheet_import_optin() -> None:
    with tempfile.TemporaryDirectory(prefix="label-rate-sheet-import-smoke-") as tmp:
        tmp_path = Path(tmp)
        source_path = tmp_path / "source.jsonl"
        write_source(source_path)
        calls: list[list[str]] = []
        expected_url = "https://bytedance.larkoffice.com/sheets/smoke-auto-import"

        original_run_lark_cli = sheet_importer.run_lark_cli

        def fake_run_lark_cli(
            command: list[str],
            *,
            cwd: Path | None = None,
        ) -> dict[str, Any]:
            del cwd
            calls.append(command)
            return {"data": {"url": expected_url}}

        sheet_importer.run_lark_cli = fake_run_lark_cli
        try:
            # Default (opt-out): no sheet_url and no opt-in must NOT import.
            default_output_dir = tmp_path / "output_default"
            default_artifacts = (
                notification_artifacts.build_label_rate_notification_artifacts(
                    source_path=source_path,
                    output_dir=default_output_dir,
                    top_n=2,
                    sheet_url=None,
                    identity="bot",
                    title="近7天低效打标策略全等级结果",
                    self_send_requested=False,
                    sent_payload=None,
                    target_user_id=None,
                    target_chat_id=None,
                )
            )
            if calls:
                raise AssertionError(
                    "sheet import must not run without explicit opt-in."
                )
            if default_artifacts.publish_summary.get("sheet_url") not in (None, ""):
                raise AssertionError(
                    "default run must not fill an online sheet_url."
                )
            if (default_output_dir / "sheet_import_result.json").exists():
                raise AssertionError(
                    "default run must not write a sheet import result."
                )

            # Opt-in: auto_import_sheet=True must import and fill the url.
            optin_output_dir = tmp_path / "output_optin"
            artifacts = (
                notification_artifacts.build_label_rate_notification_artifacts(
                    source_path=source_path,
                    output_dir=optin_output_dir,
                    top_n=2,
                    sheet_url=None,
                    identity="bot",
                    title="近7天低效打标策略全等级结果",
                    self_send_requested=False,
                    sent_payload=None,
                    target_user_id=None,
                    target_chat_id=None,
                    auto_import_sheet=True,
                )
            )
        finally:
            sheet_importer.run_lark_cli = original_run_lark_cli

        if not calls:
            raise AssertionError("sheet import was not attempted when opted in.")
        if "+workbook-import" not in calls[0]:
            raise AssertionError("sheet import command mismatch.")
        if artifacts.summary.get("sheet_url") != expected_url:
            raise AssertionError("summary sheet_url was not filled by opt-in import.")
        if artifacts.publish_summary.get("sheet_url") != expected_url:
            raise AssertionError(
                "publish_summary sheet_url was not filled by opt-in import."
            )
        if artifacts.send_plan.get("online_write_executed") is not True:
            raise AssertionError("send_plan must audit successful online write.")
        if (
            artifacts.notification_draft.get("send_safety", {}).get(
                "online_write_executed"
            )
            is not True
        ):
            raise AssertionError(
                "notification_draft must audit successful online write."
            )
        import_result = read_json(optin_output_dir / "sheet_import_result.json")
        if import_result.get("status") != "success":
            raise AssertionError("sheet import result must record success.")


def validate_weekly_filtered_summary_comparison() -> None:
    with tempfile.TemporaryDirectory(prefix="label-rate-weekly-comparison-") as tmp:
        tmp_path = Path(tmp)
        previous_path = tmp_path / "previous" / "汇总统计_剔除+1同意.csv"
        current_path = tmp_path / "current" / "汇总统计_剔除+1同意.csv"
        header = (
            "机审一级标签,POC,低效策略数,低效策略日均进审量,"
            "低效策略日均完审量,低效策略日均打标量,低效策略打标率\n"
        )
        previous_path.parent.mkdir(parents=True)
        current_path.parent.mkdir(parents=True)
        previous_path.write_text(
            "\ufeff" + header
            + "国家安全,杜衡,2,100,100,5,0.05\n"
            + "偏激社会情绪和涉外言论,张发奇,1,30,30,3,0.10\n",
            encoding="utf-8",
        )
        current_path.write_text(
            "\ufeff" + header
            + "国家安全,杜衡,1,120,120,12,0.10\n"
            + "领导人,宋诗慧,1,80,80,4,0.05\n",
            encoding="utf-8",
        )
        artifacts = build_weekly_summary_comparison(
            previous_summary_path=previous_path,
            current_summary_path=current_path,
            previous_start_date="2026-07-06",
            previous_end_date="2026-07-12",
            current_start_date="2026-07-13",
            current_end_date="2026-07-19",
            output_dir=tmp_path / "output",
        )

        if artifacts.sheet_url is not None:
            raise AssertionError("comparison must not import a Sheet by default.")
        if artifacts.online_write_attempted or artifacts.online_write_executed:
            raise AssertionError("comparison must keep online writes opt-in.")
        if len(artifacts.comparison_rows) != 3:
            raise AssertionError("comparison must retain the union of both periods.")
        if artifacts.comparison_rows[0]["avg_review_done_delta"] != 20:
            raise AssertionError("comparison delta mismatch.")
        if artifacts.comparison_rows[-1]["previous_strategy_count"] != 0:
            raise AssertionError("current-only summary key must retain previous zero.")
        totals = artifacts.totals
        if totals["previous_strategy_count"] != 3:
            raise AssertionError("comparison previous total strategy count mismatch.")
        if totals["current_strategy_count"] != 2:
            raise AssertionError("comparison current total strategy count mismatch.")
        if totals["previous_label_rate"] != 8 / 130:
            raise AssertionError("comparison previous weighted label rate mismatch.")
        if totals["current_label_rate"] != 16 / 200:
            raise AssertionError("comparison current weighted label rate mismatch.")

        workbook = load_workbook(artifacts.workbook_path, data_only=False)
        try:
            sheet = workbook[COMPARISON_SHEET_NAME]
            if sheet.freeze_panes != "C3":
                raise AssertionError("comparison workbook must freeze two headers and dimensions.")
            if "A1:A2" not in {str(item) for item in sheet.merged_cells.ranges}:
                raise AssertionError("comparison workbook missing grouped header merge.")
            if sheet["G3"].value != 20:
                raise AssertionError("comparison workbook delta cell mismatch.")
            color = sheet["G3"].fill.fgColor.rgb or ""
            if not color.endswith("F4CCCC"):
                raise AssertionError("positive comparison delta must be highlighted red.")
            if sheet["H5"].value != "/":
                raise AssertionError("zero-denominator growth must render as slash.")
            if "数据源：" not in str(sheet["A8"].value):
                raise AssertionError("comparison workbook missing source footer.")
        finally:
            workbook.close()


if __name__ == "__main__":
    main()
